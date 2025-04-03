#!/usr/bin/env python
# coding: utf-8

# In[ ]:


from IPython.display import display, Markdown
import time
import pytz
import gc
import os
import sys
import pathlib
import re
import json
import traceback
from hashlib import sha512

import sqlalchemy
from sqlalchemy.engine import *
from sqlalchemy import text
from sqlalchemy import types
from sqlalchemy import event
from sqlalchemy import Table
from sqlalchemy import Column
from sqlalchemy import MetaData
from sqlalchemy.orm import sessionmaker

from pyspark.sql import SparkSession
from pyspark.sql import SQLContext
from pyspark.sql import DataFrame
from pyspark.sql import functions as F
from pyspark.sql.types import *
from pyspark.sql import Row
from pyspark.sql import HiveContext
from pyspark.sql.functions import udf
from pyspark import SparkContext
from pyspark import SparkConf
from pyspark.sql.functions import *

import math
import random
import numpy as np
import pandas as pd

from pyexcelerate import Workbook
from openpyxl import load_workbook

import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
import seaborn as sns


current_directory = 'C:\\Users\\'+os.environ.get('USERNAME')
os.chdir(current_directory)
folder_path = current_directory+'\\Documents\\Python-Oracle\\Excel'
if os.path.exists(folder_path) == True:
    print(f'Folder path {folder_path} already exists.')
else:
    print(f'Folder path {folder_path} does not exists. Creating new directory.')
    # Create the new folder
    os.makedirs(folder_path, exist_ok=True)

def fname_check(raw_fname,table_name=None):
    from pathlib import Path
    curr_folder_path = r'C:/Users/'+os.environ.get('USERNAME')+'/Documents/Python-Oracle/Excel'
    special_characters = "!@#$%^&*()[]{};:,./<>?\|`~-=+"
    
    if 'C:/Users/' in raw_fname and '/Documents/Python-Oracle/Excel' in raw_fname:
        fname = raw_fname
        if is_nan(table_name) == True:
            filename = fname.split('/')
            filename = filename[-1].split('.')
            table_name = filename[0].upper()
        else:
            table_name = table_name.upper()
    elif '/' in raw_fname:
        filename = raw_fname.split('/')
        fname = filename[-1]
        if is_nan(table_name) == True:
            filename = fname.split('/')
            filename = filename[-1].split('.')
            table_name = filename[0].upper()
        else:
            table_name = table_name.upper()
    elif '/' not in raw_fname and raw_fname.endswith('.xlsx'):
        fname = raw_fname
        if is_nan(table_name) == True:
            filename = fname.split('/')
            filename = filename[-1].split('.')
            table_name = filename[0].upper()
        else:
            table_name = table_name.upper()
    else:
        filename = raw_fname.split('.')
        filename = filename[-2:]
        fname = '.'.join(filename)
        if is_nan(table_name) == True:
            table_name = filename[0].upper()
        else:
            table_name = table_name.upper()
    
    filename = fname.split('/')
    fname_clean = filename[-1]
    filename_clean = fname_clean.split('.')
    if any(item in special_characters for item in filename_clean[0]):
        fname_clean = clean_char(filename_clean[0],"_") # replace special characters with "_" and remove ending "_" character.
        fname_clean = filename[:-1]+fname_clean+'.'+filename_clean[-1]
        fname = fname_clean
        
    fullpath = None
    is_fullpath = False
    is_file = False
    folder_path = curr_folder_path
    
    if os.path.exists(curr_folder_path+'/'+fname) == True:
        fullpath = str(Path(curr_folder_path+'/'+fname).resolve())
        is_fullpath = os.path.exists(curr_folder_path+'/'+fname)
        is_file = os.path.isfile(fullpath)
        folder_path = fullpath.split('\\')
        folder_path = '/'.join(folder_path[:-1])
        fname = fullpath.replace('\\','/')
    else:
        fname = raw_fname
        if '/' in fname:
            filename = fname.split('/')
            filename = filename[-1].split('.')
            # Convert to uppercase for case sensitivity
            fname = filename[0].upper()
            fname_clean = clean_char(filename[0],"_") # replace special characters with "_" and remove ending "_" character.
        else:
            filename = fname.split('.')
            # Convert to uppercase for case sensitivity
            fname = filename[0].upper()
            fname_clean = clean_char(filename[0],"_") # replace special characters with "_" and remove ending "_" character.

        for i in range(1,len(filename)):
            fname += '.' + filename[i]
            fname_clean += '.' + filename[i]

        dataset_dir = pathlib.Path(curr_folder_path)
        dataset_dir_list = [item for item in dataset_dir.rglob("*") if item.is_dir()]
        dataset_file_list = [item for item in dataset_dir.rglob("*") if item.is_file()]
        dataset_dir_array_shape = np.array(dataset_dir_list).shape
        dataset_file_array_shape = np.array(dataset_file_list).shape
        for j in range(dataset_file_array_shape[0]):
            pathlib_fullpath = str(dataset_file_list[j])
            # If fname contains special characters, use fname_clean
            if any(item in special_characters for item in fname):
                os_fullpath = os.path.join(folder_path,fname_clean)
                if os_fullpath == pathlib_fullpath:
                    fullpath = os.path.join(folder_path,fname)
                    is_fullpath = os.path.exists(fullpath)
                    is_file = os.path.isfile(fullpath)
                    fname = folder_path.replace("\\","/")+'/'+fname_clean
            else:
                os_fullpath = os.path.join(folder_path,fname)
                if os_fullpath == pathlib_fullpath:
                    fullpath = os.path.join(folder_path,fname)
                    is_fullpath = os.path.exists(fullpath)
                    is_file = os.path.isfile(fullpath)
                    fname = folder_path.replace("\\","/")+'/'+fname
        
                        
    _dict_ = {
        'is_fullpath': is_fullpath,
        'fullpath': fullpath,
        'is_file': is_file,
        'folder_path': folder_path,
        'fname': fname,
        'table_name': table_name
    }
    return _dict_

def get_credentials(filename,schema=None,database='oracle',db_file=None):
    current_directory = 'C:\\Users\\'+os.environ.get('USERNAME')
    filename = current_directory+'\\'+filename
    credential = open(filename,'r')
    USERNAME = credential.readline().rstrip()
    PASSWORD = credential.readline().rstrip()
    OS_USER = credential.readline().rstrip()
    HOST = credential.readline().rstrip()
    PORT = credential.readline().rstrip()
    SERVICE = credential.readline()
    if database.lower() == 'oracle':
        if schema != None:
            USERNAME = USERNAME+f'[{schema.upper()}]'
        else:
            USERNAME = USERNAME
        ENGINE_PATH_WIN_AUTH = f'oracle+cx_oracle://{USERNAME}:{PASSWORD}@{HOST}:{PORT}/?service_name={SERVICE}'
    elif database.lower() == 'mysql':
        if db_file != None:
            ENGINE_PATH_WIN_AUTH = f'mysql+pymysql://{USERNAME}:{PASSWORD}@{HOST}:{PORT}/{db_file}'
        else:
            ENGINE_PATH_WIN_AUTH = f'mysql+pymysql://{USERNAME}:{PASSWORD}@{HOST}:{PORT}'
    elif database.lower() == 'sqlite':
        if db_file != None:
            db_fullpath = fname_check(db_file)
        else:
            db_fullpath = input('Missing db_file argument. Insert database file path.')
        ENGINE_PATH_WIN_AUTH = f'sqlite://{db_fullpath}'
    elif database.lower() == 'postgresql':
        if db_file != None:
            ENGINE_PATH_WIN_AUTH = f'postgresql://{USERNAME}:{PASSWORD}@{HOST}:{PORT}/{db_file}'
        else:
            ENGINE_PATH_WIN_AUTH = f'postgresql://{USERNAME}:{PASSWORD}@{HOST}:{PORT}'
    credential.close()
    
    _dict = {
        'DB_USERNAME': USERNAME,
        'DB_PASSWORD': PASSWORD,
        'HOST': HOST,
        'PORT': PORT,
        'SERVICE': SERVICE,
        'OS_USERNAME': OS_USER,
        'ENGINE_PATH_WIN_AUTH': ENGINE_PATH_WIN_AUTH,
        'DATABASE': database
    }
    return _dict

def connect_sqlalchemy(credentials):
    from sqlalchemy import create_engine
    ENGINE = create_engine(credentials['ENGINE_PATH_WIN_AUTH'], pool_pre_ping=True)
    return ENGINE

#--- Ths is the workaround for fast_executemany to speed-up INSERT.
# @event.listens_for(engine, 'before_cursor_execute')
def receive_before_cursor_execute(conn, cursor, statement, params, context, executemany):
    if executemany:
        cursor.fast_executemany = True

def get_time_elapsed(start_time):
    import time
    print(f'Time elapsed: {time.time() - start_time:2f}')
    new_start_time = time.time()
    return new_start_time
        
def replace_special_char(string,character):
    special_characters = "!@#$%^&*()[]{};:,./<>?\|`~-=+"
    if any(item in special_characters for item in string):
        string = string.translate ({ord(c): character for c in special_characters}) # excluding defined character
    return string

def remove_digits(string):
    string = ''.join([i for i in string if not i.isdigit()])
    return string

def remove_end_char(string,character):
    if string.endswith(character):
        string = string[:-1]
    return string

def clean_char(string,character):
    string = replace_special_char(string,character)
    string = remove_end_char(string,character)
    return string

def insert_space_at_line_breaks(text,num_indent):
    lines = text.split('\n')
    lines_with_space = [' '*num_indent + line for line in lines]
    return '\n'.join(lines_with_space)

def days_between(d1,d2,temporal='days'):
    import datetime
    from dateutil.relativedelta import relativedelta
    from datetime import date
    d1_split = d1.split("'")
    if d1_split[-1] == '':
        d1_split.pop(-1)
    d2_split = d2.split("'")
    if d2_split[-1] == '':
        d2_split.pop(-1)
    if "DATE '" in d1 and '-' in d1 and len(d1_split) > 2:
        d1 = "'".join(d1_split[:2])+"'"
        d1_inc = d1_split[-1].strip().split(' ')
        inc = []
        for i in d1_inc:
            if len(d1_inc) == 1:
                d1_inc = int(i)
            else:
                if i == "-":
                    inc.append(i)
                elif i == "+":
                    pass
                elif i == "":
                    pass
                elif type(int(i)) == int:
                    inc.append(i)
        d1_inc = int(''.join(inc))
        d1 = ora_date_to_pydate(d1)
        d1 += relativedelta(days=d1_inc)
    elif "DATE '" in d1 and '-' in d1:
        d1 = ora_date_to_pydate(d1)
    elif 'TRUNC(SYSDATE' in d1:
        if '-' in d1:
            d1_inc = d1.split('-')[-1].strip()
            d1_inc = -int(d1_inc)
            d1 = date.today()
            d1 += relativedelta(days=d1_inc)
            d1 = 'DATE ' + "'" + str(d1) + "'"
        elif '+' in d1:
            d1_inc = d1.split('+')[-1].strip()
            d1_inc = int(d1_inc)
            d1 = date.today()
            d1 += relativedelta(days=d1_inc)
            d1 = 'DATE ' + "'" + str(d1) + "'"
        else:
            d1 = date.today()
            d1 = 'DATE ' + "'" + str(d1) + "'"
        d1 = ora_date_to_pydate(d1)
    else:
        d1 = datetime.datetime.strptime(d1, "%Y-%m-%d")
            
    if "DATE '" in d2 and '-' in d2 and len(d2_split) > 2:
        d2 = "'".join(d2_split[:2])+"'"
        d2_inc = d2_split[-1].strip().split(' ')
        print(d2_split)
        print(d2_inc)
        inc = []
        for i in d2_inc:
            if len(d2_inc) == 1:
                d2_inc = int(i)
            else:
                if i == "-":
                    inc.append(i)
                elif i == "+":
                    pass
                elif i == "":
                    pass
                elif type(int(i)) == int:
                    inc.append(i)
        d2_inc = int(''.join(inc))
        d2 = ora_date_to_pydate(d2)
        d2 += relativedelta(days=d2_inc)
    elif "DATE '" in d2 and '-' in d2:
        d2 = ora_date_to_pydate(d2)
    elif 'TRUNC(SYSDATE' in d2:
        if '-' in d2:
            d2_inc = d2.split('-')[-1].strip()
            d2_inc = -int(d2_inc)
            d2 = date.today()
            d2 += relativedelta(days=d2_inc)
            d2 = 'DATE ' + "'" + str(d2) + "'"
        elif '+' in d2:
            d2_inc = d2.split('+')[-1].strip()
            d2_inc = int(d2_inc)
            d2 = date.today()
            d2 += relativedelta(days=d2_inc)
            d2 = 'DATE ' + "'" + str(d2) + "'"
        else:
            d2 = date.today()
            d2 = 'DATE ' + "'" + str(d2) + "'"
        d2 = ora_date_to_pydate(d2)
    else:
        d2 = datetime.datetime.strptime(d2, "%Y-%m-%d")
        
    if temporal.upper() == 'DAYS' or temporal.upper() == 'DAY' or temporal.upper() == 'DAILY' or temporal.upper() == 'DD':
        diff = abs((d2-d1).days)
    elif temporal.upper() == 'WEEKS' or temporal.upper() == 'WEEK' or temporal.upper() == 'WEEKLY' or temporal.upper() == 'WW' or temporal.upper() == 'IW':
        diff = abs((d2-d1).days // 7) + 1
    elif temporal.upper() == 'MONTHS' or temporal.upper() == 'MONTH' or temporal.upper() == 'MONTHLY' or temporal.upper() == 'MM':
        diff = abs(relativedelta(d2, d1).years)*12 + abs(relativedelta(d2, d1).months) + 1
    elif temporal.upper() == 'YEARS' or temporal.upper() == 'YEAR' or temporal.upper() == 'YEARLY' or temporal.upper() == 'YY'  or temporal.upper() == 'YYYY':
        diff = abs(relativedelta(d2, d1).years) + 1
        
    return diff

def ora_date_to_pydate(var_date):
    from datetime import date
    from dateutil.relativedelta import relativedelta
    if 'TRUNC(SYSDATE' in var_date:
        if '-' in var_date:
            var_date_inc = var_date.split('-')[-1].strip()
            var_date_inc = -int(var_date_inc)
            var_date = date.today()
            var_date += relativedelta(days=var_date_inc)
            var_date = 'DATE ' + "'" + str(var_date) + "'"
        elif '+' in var_date:
            var_date_inc = var_date.split('+')[-1].strip()
            var_date_inc = int(var_date_inc)
            var_date = date.today()
            var_date += relativedelta(days=var_date_inc)
            var_date = 'DATE ' + "'" + str(var_date) + "'"
        else:
            var_date = date.today()
            var_date = 'DATE ' + "'" + str(var_date) + "'"
    var_date_result = re.search(" '"+"(.*)"+"'",var_date)
    var_date_year = var_date_result.group(1)[0:4]
    var_date_month = var_date_result.group(1)[5:7]
    var_date_day = var_date_result.group(1)[8:10]
    pydate = date(int(var_date_year), int(var_date_month), int(var_date_day))
    return pydate

def ora_date_inc(var_date,temporal,increment):
    from datetime import date
    from dateutil.relativedelta import relativedelta
    if 'TRUNC(SYSDATE' in var_date:
        if '-' in var_date:
            var_date_inc = var_date.split('-')[-1].strip()
            var_date_inc = -int(var_date_inc)
            var_date = date.today()
            var_date += relativedelta(days=var_date_inc)
            var_date = 'DATE ' + "'" + str(var_date) + "'"
        elif '+' in var_date:
            var_date_inc = var_date.split('+')[-1].strip()
            var_date_inc = int(var_date_inc)
            var_date = date.today()
            var_date += relativedelta(days=var_date_inc)
            var_date = 'DATE ' + "'" + str(var_date) + "'"
        else:
            var_date = date.today()
            var_date = 'DATE ' + "'" + str(var_date) + "'"
    var_date = ora_date_to_pydate(var_date)
    if temporal.upper() == 'DAYS' or temporal.upper() == 'DAY' or temporal.upper() == 'DAILY' or temporal.upper() == 'DD':
        var_date += relativedelta(days=increment)
    elif temporal.upper() == 'WEEKS' or temporal.upper() == 'WEEK' or temporal.upper() == 'WEEKLY' or temporal.upper() == 'WW' or temporal.upper() == 'IW':
        var_date += relativedelta(weeks=increment)
    elif temporal.upper() == 'MONTHS' or temporal.upper() == 'MONTH' or temporal.upper() == 'MONTHLY' or temporal.upper() == 'MM':
        var_date += relativedelta(months=increment)
    elif temporal.upper() == 'YEARS' or temporal.upper() == 'YEAR' or temporal.upper() == 'YEARLY' or temporal.upper() == 'YY'  or temporal.upper() == 'YYYY':
        var_date += relativedelta(years=increment)
    date_string = 'DATE ' + "'" + str(var_date) + "'"
    return date_string

def df_fix_dtype(df,schema,dtyp=None):
    import datetime
    from datetime import date
    import decimal
    from decimal import Decimal
    
    df_sample = df.loc[:10000,:]
    col_w_nan = df_sample.columns[df_sample.isna().any()].tolist()
    col_n_nan = df_sample.columns[df_sample.notnull().all()]
    
    df_col_w_nan_new = pd.DataFrame()
    df_col_w_nan_new[col_w_nan] = np.nan
    
    for cols,rows in df[col_w_nan].items():
        row_sample = rows.loc[:10000]
        is_int_type = row_sample[row_sample.notnull()].apply(lambda x: isinstance(x,int)).any()
        is_float_type = row_sample[row_sample.notnull()].apply(lambda x: isinstance(x,float)).any()
        is_date_type = check_dtime(row_sample)['is_date']
        is_datetime_type = check_dtime(row_sample)['is_timestamp']
        is_str_type = row_sample[row_sample.notnull()].apply(lambda x: isinstance(x,str)).any()
        
        if is_float_type == True:
            df_col_w_nan_new[cols] = pd.to_numeric(rows,errors='coerce')
        elif is_int_type == True:
            df_col_w_nan_new[cols] = pd.to_numeric(rows,errors='coerce').astype(pd.Int64Dtype())
        elif is_date_type == True or is_datetime_type == True:
            df_col_w_nan_new[cols] = pd.to_datetime(rows,errors='coerce').dt.strftime('%Y-%m-%d %H:%M:%S').astype(str)
        elif is_str_type == True:
            df_col_w_nan_new[cols] = rows.astype(str)
        
    for cols in df_sample[col_n_nan].columns:
        if df_sample[cols].apply(lambda x: isinstance(x,datetime.date)).any() == True:
            df[cols] = pd.to_datetime(df[cols], errors='coerce').dt.strftime('%Y-%m-%d %H:%M:%S').astype(str)
    
    my_credentials = get_credentials('my_credentials.txt',schema)
    try:
        engine = connect_sqlalchemy(my_credentials)
        if dtyp == None:
            df_ora_dtyp = get_sqlalchemy_dtyp_dict(df,engine)
        else:
            df_ora_dtyp = dtyp
        engine.dispose()
    except Exception as err:
        print('Error: ' + str(err))
        engine.dispose()
    else:
        col_nan_all = df_sample[col_w_nan].loc[:,~df_sample[col_w_nan].columns.isin(df_col_w_nan_new)].columns
        keys, values = zip(*df_ora_dtyp.items())
        list_key = list(keys)
        list_val = list(values)
        for cols in col_nan_all:
            list_key_idx = list_key.index(cols)
            try:
                sql2py = list_val[list_key_idx].python_type
            except:
                sql2py = str
            if sql2py == datetime.datetime:
                df_col_w_nan_new[cols] = pd.to_datetime(df[cols],errors='coerce')
            elif sql2py == str:
                df_col_w_nan_new[cols] = df[cols].astype(str)
            elif sql2py == int:
                df_col_w_nan_new[cols] = pd.to_numeric(df[cols],errors='coerce').astype(pd.Int64Dtype())
            elif sql2py == float:
                df_col_w_nan_new[cols] = pd.to_numeric(df[cols],errors='coerce')
            elif sql2py == decimal.Decimal:
                df_col_w_nan_new[cols] = pd.to_numeric(df[cols],errors='coerce')
            else:
                df_col_w_nan_new[cols] = df[cols]

        df[col_w_nan] = df_col_w_nan_new
        del df_col_w_nan_new
    return df

def query_to_df(query,schema,chunk_size=1000000):
    # Open
    my_credentials = get_credentials('my_credentials.txt',schema)
    engine = connect_sqlalchemy(my_credentials)
    try:
        start_time = time.time()
        if chunk_size != None:
            # Create empty list
            dfl = []
            # Create empty dataframe
            dfs = pd.DataFrame()
            # Start Chunking SQL
            chunk_idx = 0
            start_time_file = time.time()
            start_time_chunk = time.time()
            for chunk in pd.read_sql_query(query,engine,chunksize=chunk_size):
                # Start Appending Data Chunks from SQL Result set into List
                dfl.append(chunk)
                chunk_idx += 1
                end_time_chunk = time.time()
                print('CHUNK {} appended to list. Time elapsed: {:2f}'.format(chunk_idx,end_time_chunk - start_time_chunk))
                start_time_chunk = time.time()
            end_time_file = time.time()
            print('Chunking done. Time elapsed: {:2f}'.format(end_time_file - start_time_file),'s')
            # Append data from list to dataframe
            start_time_df = time.time()
            dfs = pd.concat(dfl, ignore_index=True)
            dfs.columns = dfs.columns.str.upper()
            dfs.where(pd.notnull(dfs),np.nan,inplace=True)
            dfs = df_fix_dtype(dfs,schema)
            del dfl # delete variable to clear memory cache
            print('Chunk list deleted.')
            end_time_df = time.time()
            print('Dataframe created. Time elapsed: {:2f}'.format(end_time_df - start_time_df),'s')
        else:
            start_time_df = time.time()
            dfs = pd.read_sql_query(query,engine)
            dfs.columns = dfs.columns.str.upper()
            dfs.where(pd.notnull(dfs),np.nan,inplace=True)
            dfs = df_fix_dtype(dfs,schema)
            end_time_df = time.time()
            print('Dataframe created. Time elapsed: {:2f}'.format(end_time_df - start_time_df),'s')

        end_time = time.time()
        print('Execution time: {:2f}'.format(end_time - start_time),'s')
        engine.dispose()
    except Exception as err:
        print('Error: ' + str(err))
        engine.dispose()
    return dfs

def query_to_df_whole_timeline(query,schema,date_start,date_end,chunk_size=1000000):
    start_time = time.time()
    my_credentials = get_credentials('my_credentials.txt',schema)
    engine = connect_sqlalchemy(my_credentials)
    try:
        if chunk_size != None:
            # Create empty list
            dfl = []
            # Create empty dataframe
            dfs = pd.DataFrame()
            # Start Chunking SQL
            chunk_idx = 0
            start_time_file = time.time()
            start_time_chunk = time.time()
            for chunk in pd.read_sql_query(query,engine,chunksize=chunk_size):
                # Start Appending Data Chunks from SQL Result set into List
                dfl.append(chunk)
                chunk_idx += 1
                end_time_chunk = time.time()
                exec_time = end_time_chunk - start_time_chunk
                print('CHUNK {} appended to list. Time elapsed: {:2f}'.format(chunk_idx,exec_time))
                start_time_chunk = time.time()

            end_time_file = time.time()
            print('Chunking done. Time elapsed: {:2f}'.format(end_time_file - start_time_file),'s')
            # Append data from list to dataframe
            start_time_df = time.time()
            dfs = pd.concat(dfl, ignore_index=True)
            dfs.columns = dfs.columns.str.upper()
            dfs.where(pd.notnull(dfs),np.nan,inplace=True)
            dfs = df_fix_dtype(dfs,schema)
            is_df_created = 'YES'
            del dfl # delete variable to clear memory cache
            print('Chunk list deleted.')
            end_time_df = time.time()
            print('Dataframe created. Time elapsed: {:2f}'.format(end_time_df - start_time_df),'s')
        else:
            start_time_df = time.time()
            dfs = pd.read_sql_query(query,engine)
            dfs.columns = dfs.columns.str.upper()
            dfs.where(pd.notnull(dfs),np.nan,inplace=True)
            dfs = df_fix_dtype(dfs,schema)
            end_time_df = time.time()
            print('Dataframe created. Time elapsed: {:2f}'.format(end_time_df - start_time_df),'s')

        end_time = time.time()
        exec_time = end_time - start_time
        print('Execution time: {:2f}'.format(exec_time),'s')
        engine.dispose()
    except Exception as err:
        print('Error:' + str(err))
        engine.dispose()
        
    return dfs

def replace_word_containing_substring(text, substring, replacement):
    # Define the regex pattern to match whole words containing the substring
    pattern = r'\b\w*' + re.escape(substring.lower()) + r'\w*\b'
    # Use re.sub to replace the matched words with the replacement
    result = re.sub(pattern, replacement, text)
    return result

def var_query_fstring(text,c_date_start,c_date_end):
    query = replace_word_containing_substring(text, 'date_start', c_date_start)
    query = replace_word_containing_substring(query, 'date_end', c_date_end)
    return f'''{query}'''

def var_query_to_df(query_c_date,schema,date_start,date_end,time_out=60,temporal='days',increment=1,chunk_size = 1000000):
    # Open
    my_credentials = get_credentials('my_credentials.txt',schema)
    c_date_start = date_start
    c_date_end = date_end
    query = var_query_fstring(query_c_date,c_date_start,c_date_end)
    date_diff = days_between(c_date_start,c_date_end,temporal)
    engine = connect_sqlalchemy(my_credentials)
    try:
        print('Timeline: {} to {}'.format(c_date_start,c_date_end))
        start_time = time.time()
        # Create empty list
        dfl = []
        # Create empty dataframe
        dfs = pd.DataFrame()
        # Start Chunking SQL
        chunk_idx = 0
        start_time_file = time.time()
        start_time_chunk = time.time()
        
        from func_timeout import func_timeout, FunctionTimedOut
        try:
            dfs = func_timeout(time_out, query_to_df_whole_timeline, args=(query,schema,date_start,date_end,chunk_size))
        except FunctionTimedOut:
            if temporal.upper() == 'DAYS' or temporal.upper() == 'DAY' or temporal.upper() == 'DAILY' or temporal.upper() == 'DD':
                if increment == 1:
                    print(f'Process takes too long. Partitioning timeline by day and populating individually.')
                else:
                    print(f'Process takes too long. Partitioning timeline by {increment} days and populating individually.')
            elif temporal.upper() == 'WEEKS' or temporal.upper() == 'WEEK' or temporal.upper() == 'WEEKLY' or temporal.upper() == 'WW' or temporal.upper() == 'IW':
                if increment == 1:
                    print(f'Process takes too long. Partitioning timeline by week and populating individually.')
                else:
                    print(f'Process takes too long. Partitioning timeline by {increment} weeks and populating individually.')
            elif temporal.upper() == 'MONTHS' or temporal.upper() == 'MONTH' or temporal.upper() == 'MONTHLY' or temporal.upper() == 'MM':
                if increment == 1:
                    print(f'Process takes too long. Partitioning timeline by month and populating individually.')
                else:
                    print(f'Process takes too long. Partitioning timeline by {increment} months and populating individually.')
            elif temporal.upper() == 'YEARS' or temporal.upper() == 'YEAR' or temporal.upper() == 'YEARLY' or temporal.upper() == 'YY'  or temporal.upper() == 'YYYY':
                if increment == 1:
                    print(f'Process takes too long. Partitioning timeline by year and populating individually.')
                else:
                    print(f'Process takes too long. Partitioning timeline by {increment} years and populating individually.')
        except Exception as err:
            print('Error: ' + str(err))
            engine.dispose()
                
        if dfs.empty == True:
            var_start_time_df = time.time()
            var_dfs = []
            c_date_end = ora_date_inc(c_date_start,temporal,increment)
            for i in range(date_diff):
                print('Timeline: {} to {}'.format(c_date_start,c_date_end))
                query = var_query_fstring(query_c_date,c_date_start,c_date_end)
                if chunk_size != None:
                    # Create empty list
                    dfl = []
                    # Create empty dataframe
                    dfs = pd.DataFrame()
                    # Start Chunking SQL
                    chunk_idx = 0
                    start_time_file = time.time()
                    start_time_chunk = time.time()
                    for chunk in pd.read_sql_query(query,engine,chunksize=chunk_size):
                        # Start Appending Data Chunks from SQL Result set into List
                        dfl.append(chunk)
                        chunk_idx += 1
                        end_time_chunk = time.time()
                        print('CHUNK {} appended to list. Time elapsed: {:2f}'.format(chunk_idx,end_time_chunk - start_time_chunk))
                        start_time_chunk = time.time()
                    end_time_file = time.time()
                    print('Chunking done. Time elapsed: {:2f}'.format(end_time_file - start_time_file),'s')
                    # Append data from list to dataframe
                    start_time_df = time.time()
                    dfs = pd.concat(dfl, ignore_index=True)
                    dfs.columns = dfs.columns.str.upper()
                    dfs.where(pd.notnull(dfs),np.nan,inplace=True)
                    dfs = df_fix_dtype(dfs,schema)
                    del dfl # delete variable to clear memory cache
                    print('Chunk list deleted.')
                    end_time_df = time.time()
                    print('Partial Dataframe {} created. Time elapsed: {:2f}'.format(i+1,end_time_df - start_time_df),'s')
                else:
                    start_time_df = time.time()
                    dfs = pd.concat(dfl, ignore_index=True)
                    dfs.columns = dfs.columns.str.upper()
                    dfs.where(pd.notnull(dfs),np.nan,inplace=True)
                    dfs = df_fix_dtype(dfs,schema)
                    del dfl # delete variable to clear memory cache
                    print('Chunk list deleted.')
                    end_time_df = time.time()
                    print('Partial Dataframe {} created. Time elapsed: {:2f}'.format(i+1,end_time_df - start_time_df),'s')
                # Start Appending Data Chunks from SQL Result set into List
                var_dfs.append(dfs)
                del dfs # delete variable to clear memory cache
                end_time = time.time()
                print('Appended to main dataframe. Time elapsed: {:2f}'.format(end_time - start_time),'s')
                c_date_start = ora_date_inc(c_date_start,temporal,increment)
                c_date_end = ora_date_inc(c_date_end,temporal,increment)
                if ora_date_to_pydate(c_date_end) > ora_date_to_pydate(date_end):
                    c_date_end = date_end
                    if c_date_end == c_date_start:
                        c_date_end = ora_date_inc(c_date_end,'days',1)
            var_end_time_df = time.time()
            print('Dataframe created. Time elapsed: {:2f}'.format(var_end_time_df - var_start_time_df),'s')
            dfs = pd.concat(var_dfs, ignore_index=True)
    except Exception as err:
        print('Error: ' + str(err))
        engine.dispose()
    return dfs

def query_create_table(query,schema,table_name,source_owner='TEST_OWNER',source_table='TEST_TABLE'):
    my_credentials = get_credentials('my_credentials.txt',schema)
    engine = connect_sqlalchemy(my_credentials)
    df = query_to_df(query,schema)
    df_col = df.columns
    df_col_typ = df.dtypes
    df_col_list = [df_col[i] for i in range(df.shape[1])]
    df_col_str = str(df_col_list)
    df_col_str = df_col_str.replace("[","")
    df_col_str = df_col_str.replace("]","")
    df_col_str = df_col_str.replace(", ",", \n")
    col_dtyp_list = []
    try:
        df_col_oracle_typ = get_oracle_data_type(df_col_str,engine,source_owner=source_owner,source_table=source_table)
        engine.dispose()
    except:
        engine.dispose()
    for i in range(len(df_col)):
        if 'date' in str(df_col_typ[i]).lower() or 'dtime' in str(df_col[i]).lower():
            col_dtyp_list.append(df_col[i] + ' ' + 'DATE')
        elif df_col[i] in df_col_oracle_typ['COLUMN_NAME'].values:
            get_index = df_col_oracle_typ.index[df_col_oracle_typ['COLUMN_NAME']==df_col[i]]
            oracle_dtyp = str(df_col_oracle_typ['DATA_TYPE'][get_index].values)
            oracle_dtyp = oracle_dtyp.replace("[","")
            oracle_dtyp = oracle_dtyp.replace("]","")
            col_dtyp_list.append(df_col[i] + ' ' + oracle_dtyp)
        elif 'int' in str(df_col_typ[i]) or 'float' in str(df_col_typ[i]) or 'num' in str(df_col_typ[i]):
            col_dtyp_list.append(df_col[i] + ' ' + 'NUMBER')
        else:
            col_dtyp_list.append(df_col[i] + ' ' + 'VARCHAR2(80 CHAR)')

    col_dtyp = str(col_dtyp_list).replace("'","")
    col_dtyp = col_dtyp.replace('"',"")
    col_dtyp = col_dtyp.replace("[","")
    col_dtyp = col_dtyp.replace("]","")
    col_dtyp = col_dtyp.replace(", ",", \n")
    query = f'''
CREATE TABLE {schema.upper()}.{table_name.upper()}
(
{col_dtyp}
)
COMPRESS FOR ALL OPERATIONS
'''
    is_partition = input('Do you want to create a partition by range? [YES/NO]  |  ')
    if is_partition.upper() == 'YES' or is_partition.upper() == 'Y':
        date_reference = input('Input column for date_reference.  |  ')
        query = f'''
{query}
PARTITION BY RANGE ({date_reference.upper()})
INTERVAL (NUMTOYMINTERVAL(1, 'MONTH'))
(
PARTITION P0 VALUES LESS THAN (TO_DATE(' 2015-01-01 00:00:00', 'SYYYY-MM-DD HH24:MI:SS'))
)
'''
    else:
        query
    return f'''{query};'''

def query_create_gtt(query,schema,table_name,source_owner='TEST_OWNER',source_table='TEST_TABLE'):
    my_credentials = get_credentials('my_credentials.txt',schema)
    engine = connect_sqlalchemy(my_credentials)
    df = query_to_df(query,schema)
    df_col = df.columns
    df_col_typ = df.dtypes
    df_col_list = [df_col[i] for i in range(df.shape[1])]
    df_col_str = str(df_col_list)
    df_col_str = df_col_str.replace("[","")
    df_col_str = df_col_str.replace("]","")
    df_col_str = df_col_str.replace(", ",", \n")
    col_dtyp_list = []
    try:
        df_col_oracle_typ = get_oracle_data_type(df_col_str,engine,source_owner=source_owner,source_table=source_table)
        engine.dispose()
    except:
        engine.dispose()
    for i in range(len(df_col)):
        if 'date' in str(df_col_typ[i]).lower() or 'dtime' in str(df_col[i]).lower():
            col_dtyp_list.append(df_col[i] + ' ' + 'DATE')
        elif df_col[i] in df_col_oracle_typ['COLUMN_NAME'].values:
            get_index = df_col_oracle_typ.index[df_col_oracle_typ['COLUMN_NAME']==df_col[i]]
            oracle_dtyp = str(df_col_oracle_typ['DATA_TYPE'][get_index].values)
            oracle_dtyp = oracle_dtyp.replace("[","")
            oracle_dtyp = oracle_dtyp.replace("]","")
            col_dtyp_list.append(df_col[i] + ' ' + oracle_dtyp)
        elif 'int' in str(df_col_typ[i]) or 'float' in str(df_col_typ[i]) or 'num' in str(df_col_typ[i]):
            col_dtyp_list.append(df_col[i] + ' ' + 'NUMBER')
        else:
            col_dtyp_list.append(df_col[i] + ' ' + 'VARCHAR2(80 CHAR)')

    col_dtyp = str(col_dtyp_list).replace("'","")
    col_dtyp = col_dtyp.replace('"',"")
    col_dtyp = col_dtyp.replace("[","")
    col_dtyp = col_dtyp.replace("]","")
    col_dtyp = col_dtyp.replace(", ",", \n")
    query = f'''
CREATE TABLE {schema.upper()}.{table_name.upper()}
(
{col_dtyp}
)
ON COMMIT PRESERVE ROWS
;
'''
    return query

def insert_space_at_line_breaks_mrg(text,num_indent):
    text_split = text.split('\n')
    if len(text_split) > 1:
        text = insert_space_at_line_breaks(text,num_indent)
    else:
        text = text
    return text

def query_merge(df,query,schema,table_name,connection_key):
    def ora_conn_key(connection_key):
        if type(connection_key) == list and len(connection_key) > 1:
            new_connection_key = None
            for idx,conn_key in enumerate(connection_key):
                connection_key[idx] = conn_key.upper()
                if idx == 0:
                    new_connection_key = '\n' + 'DM1.'+conn_key.upper() + ' = DM2.'+conn_key.upper()
                elif conn_key.upper() == connection_key[-1]:
                    new_connection_key = new_connection_key + ' and \n' + 'DM1.'+conn_key.upper() + ' = DM2.'+conn_key.upper() + '\n'
                else:
                    new_connection_key = new_connection_key + ' and \n' + 'DM1.'+conn_key.upper() + ' = DM2.'+conn_key.upper()
        elif type(connection_key) == list and len(connection_key) == 1:
            connection_key = connection_key[0].upper()
            new_connection_key = 'DM1.'+connection_key + ' = DM2.'+connection_key
        else:
            connection_key = connection_key.upper()
            new_connection_key = 'DM1.'+connection_key + ' = DM2.'+connection_key
        return connection_key,new_connection_key
    
    df_col = df.columns.str.upper()
    conn_key = ora_conn_key(connection_key)
    
    # WHEN MATCHED THEN UPDATE SET
    df_col_list_mrg = [f"DM1.{df_col[i]} = DM2.{df_col[i]}" for i in range(df.shape[1]) if df_col[i] not in conn_key[0]]
    df_col_mrg = str(df_col_list_mrg).replace("'","")
    df_col_mrg = df_col_mrg.replace("[","")
    df_col_mrg = df_col_mrg.replace("]","")
    df_col_mrg = df_col_mrg.replace(", ",", \n")
    
    #WHEN NOT MATCHED THEN
    ##INSERT
    df_col_list_ins = ['DM1.' + df_col[i] for i in range(df.shape[1])]
    df_col_ins = str(df_col_list_ins).replace("'","")
    df_col_ins = df_col_ins.replace("[","")
    df_col_ins = df_col_ins.replace("]","")
    df_col_ins = df_col_ins.replace(", ",", \n")
    ##VALUES
    df_col_list_val = ['DM2.' + df_col[i] for i in range(df.shape[1])]
    df_col_val = str(df_col_list_val).replace("'","")
    df_col_val = df_col_val.replace("[","")
    df_col_val = df_col_val.replace("]","")
    df_col_val = df_col_val.replace(", ",", \n")
    query = f'''
MERGE INTO {schema.upper()}.{table_name.upper()} DM1
USING
(
{query}
) DM2 ON ({insert_space_at_line_breaks_mrg(conn_key[1],9)})

WHEN MATCHED THEN
UPDATE SET
{insert_space_at_line_breaks(df_col_mrg,6)}

WHEN NOT MATCHED THEN
INSERT
{' '*6}(
{insert_space_at_line_breaks(df_col_ins,6)}
{' '*6})
VALUES
{' '*6}(
{insert_space_at_line_breaks(df_col_val,6)}
{' '*6})
{' '*6};
'''
    return query

def query_create_package(query,query_sample,schema,table_name,connection_key):
    from datetime import date
    df = query_to_df(query_sample,schema)
    merge_statement = query_merge(df,query,schema,table_name,connection_key)
    
    
    query_head = f'''
CREATE OR REPLACE PACKAGE PCKG_{table_name.upper()} AS

    /****************************************************************************************************************************
    
        NAME:      PCKG_{table_name.upper()}
        PURPOSE:   Package to update {table_name.upper()}
        REVISIONS: 
        Ver        Date        Author                     Description
        ---------  ----------  ---------------            -----------------------------------------------------------------------
        1.0        {str(date.today())}  {schema.upper()}       1. Created this package.
              
    ****************************************************************************************************************************/

    procedure proc_main (p_date_start date default trunc(sysdate-1), p_date_end date default trunc(sysdate));

END PCKG_{table_name.upper()};
'''
    
    
    query_body = f'''
CREATE OR REPLACE PACKAGE BODY PCKG_{table_name.upper()} AS
    
/*----------------
------------------
----- proc00 -----
------------------
----------------*/
             
            
    procedure pstats( actable varchar2, anperc number default 0.01)
    is
    begin
        ap_public.core_log_pkg.pstart( 'Stat:'||actable );
        dbms_stats.gather_table_stats ( ownname => user, tabname => actable, estimate_percent => anperc );
        ap_public.core_log_pkg.pend;
    end pstats
    ;
    
/*----------------
------------------
----- proc01 -----
------------------
----------------*/
             
            
    procedure ptruncate( actable varchar2, anperc number default 0.01)
    is
    begin
        ap_public.core_log_pkg.pstart( 'trunc:'||actable );
        execute immediate 'truncate table '||actable ; 
        ap_public.core_log_pkg.pend ;
    end ptruncate
    ;
    
/*----------------
------------------
----- proc02 -----
------------------
----------------*/

    procedure proc_main
    (p_date_start date default trunc(sysdate-1), p_date_end date default trunc(sysdate))
    as
    --=== parametrization ===--
    
    c_date_start constant date := trunc(p_date_start);
    c_date_end constant date := trunc(p_date_end);
    
    
    begin
        ap_public.core_log_pkg.pinit( user, $$plsql_unit ) ;
    
        /*--------------------------------------------------------------------------------------------------------------------------------------------
            Procedure in updating POS EC DATAMART TREE main table
                VERSION         DATE                        USER                            NOTE
                  1         {str(date.today())}              {schema.upper()}            Created this procedure.
                  
        --------------------------------------------------------------------------------------------------------------------------------------------*/
        --ptruncate('{schema.upper()}.{table_name.upper()}');
        
        ap_public.core_log_pkg.pstart('merge: {schema.upper()}.{table_name.upper()} ' || c_date_start || ' - ' || c_date_end);
{insert_space_at_line_breaks(merge_statement,8)}
        pstats('{table_name.upper()}');
        ap_public.core_log_pkg.pfinish;
        commit;
              
        exception when others then
        ap_public.core_log_pkg.perror; 
        raise;

    end proc_main;
  
END PCKG_{table_name.upper()};
'''
    return query_head,query_body

def query_merge_when_matched_and_or_not_matched(query,schema,table_name):
    df = query_to_df(query,schema)
    df_col = df.columns
    
    # WHEN MATCHED THEN UPDATE SET
    df_col_list_mrg = ['DM1.' + df_col[i] + ' ' + '=' + ' ' + 'DM2.' + df_col[i] for i in range(df.shape[1])]
    df_col_mrg = str(df_col_list_mrg).replace("'","")
    df_col_mrg = df_col_mrg.replace("[","")
    df_col_mrg = df_col_mrg.replace("]","")
    df_col_mrg = df_col_mrg.replace(", ",", \n")
    
    #WHEN NOT MATCHED THEN
    ##INSERT
    df_col_list_ins = ['DM1.' + df_col[i] for i in range(df.shape[1])]
    df_col_ins = str(df_col_list_ins).replace("'","")
    df_col_ins = df_col_ins.replace("[","")
    df_col_ins = df_col_ins.replace("]","")
    df_col_ins = df_col_ins.replace(", ",", \n")
    ##VALUES
    df_col_list_val = ['DM2.' + df_col[i] for i in range(df.shape[1])]
    df_col_val = str(df_col_list_val).replace("'","")
    df_col_val = df_col_val.replace("[","")
    df_col_val = df_col_val.replace("]","")
    df_col_val = df_col_val.replace(", ",", \n")
    query = f'''
WHEN MATCHED THEN
UPDATE SET
{df_col_mrg}

WHEN NOT MATCHED THEN
INSERT
(
{df_col_ins}
)
VALUES
(
{df_col_val}
)
;
COMMIT;
'''
    return query

def query_insert_when_not_matched(query,schema,table_name):
    df = query_to_df(query,schema)
    df_col = df.columns
    df_col_list_ins = ['DM1.'+df_col[i] for i in range(df.shape[1])]
    df_col_ins = str(df_col_list_ins).replace("'","")
    df_col_ins = df_col_ins.replace("[","")
    df_col_ins = df_col_ins.replace("]","")
    df_col_ins = df_col_ins.replace(", ",", \n")
    df_col_list_val = ['DM2.'+df_col[i] for i in range(df.shape[1])]
    df_col_val = str(df_col_list_val).replace("'","")
    df_col_val = df_col_val.replace("[","")
    df_col_val = df_col_val.replace("]","")
    df_col_val = df_col_val.replace(", ",", \n")
    query = f'''
WHEN NOT MATCHED THEN
INSERT
(
{df_col_ins}
)
VALUES
(
{df_col_val}
)
;
'''
    return query

def get_oracle_data_type(df_col_str,engine,source_owner='TEST_OWNER',source_table='TEST_TABLE'):
    query = f'''
with tab_base as
(
select
       column_name
      ,data_type
      ,max(char_length) char_length
      ,count(*) antidup
from all_tab_columns
where 1=1
      and owner = '{source_owner.upper()}'
      and table_name = '{source_table.upper()}'
      and column_name in ({df_col_str})
group by
       column_name
      ,data_type
)
select
       column_name
      ,case
            when data_type like '%CHAR%' then data_type ||'('||char_length||' CHAR)'
            else data_type
       end                                                                      data_type
      ,char_length
from tab_base
'''
    df = pd.read_sql_query(query,engine)
    df.columns = df.columns.str.upper()
    return df

def apply_ora_hint(query):
    # Oracle only applies hint after the first select statement
    word = 'select'
    word_len = len(word)
    word_idx = query.find(word)

    hint_start = '/*+'
    hint_start_len = len(hint_start)
    hint_start_idx = query.find(hint_start)

    hint_end = '*/'
    hint_end_re = re.escape(hint_end) # workaround when keyword starts with special character
    hint_end_len = len(hint_end)
    hint_end_idxs = [i.start() for i in re.finditer(hint_end_re, query) if i.start() > hint_start_idx]
    for i in range(len(hint_end_idxs)):
        if hint_end_idxs[i] > hint_start_idx:
            idx = (np.abs(hint_end_idxs[i] - hint_start_idx)).argmin()
            hint_end_idx = hint_end_idxs[idx]

    hint = '/*+ materialize */'
    hint_len = len(hint)
    hint_idx = query.find(hint)

    if hint_start in query:
        query = query.replace(query[hint_start_idx:hint_end_idx+hint_end_len],hint)
        if query[word_len+word_idx+1:hint_len+hint_idx] == hint:
            query = query
        else:
            query = query[:word_idx] + word + ' ' + hint + query[word_len+word_idx:]
    else:
        query = query[:word_idx] + word + ' ' + hint + query[word_len+word_idx:]
    return query

def fix_ora_string(string,substr="'",new_substr="''"):
    if substr in string:
        string_list = list(string)
        substr_idxs = [i.start() for i in re.finditer(substr, string)]
        for idx in substr_idxs:
            string_list[idx] = new_substr
        string = ''.join(string_list)
    return string

def fix_ora_col_length(df1,df2,engine,cursor,schema,table_name,source_owner,source_table):
    df1_col = df1.columns
    df1_col_list = [df1_col[i] for i in range(df1.shape[1])]
    df1_col_str = str(df1_col_list)
    df1_col_str = df1_col_str.replace("[","")
    df1_col_str = df1_col_str.replace("]","")
    df1_col_str = df1_col_str.replace(", ",", \n")
    df1_dtyp = get_oracle_data_type(df1_col_str,engine,source_owner=schema,source_table=table_name)

    df2_col = df2.columns
    df2_col_list = [df2_col[i] for i in range(df2.shape[1])]
    df2_col_str = str(df2_col_list)
    df2_col_str = df2_col_str.replace("[","")
    df2_col_str = df2_col_str.replace("]","")
    df2_col_str = df2_col_str.replace(", ",", \n")
    df2_dtyp = get_oracle_data_type(df2_col_str,engine,source_owner=source_owner,source_table=source_table)

    for col in df1_col:
        idx1 = df1_dtyp[df1_dtyp['COLUMN_NAME'] == col].index[0]
        idx2 = df2_dtyp[df2_dtyp['COLUMN_NAME'] == col].index[0]
        if 'VARCHAR2' in df1_dtyp.loc[idx1,'DATA_TYPE'] and df1_dtyp.loc[idx1,'CHAR_LENGTH'] < df2_dtyp.loc[idx2,'CHAR_LENGTH']:
            print(f'''Modifying {schema}.{table_name}.{df1_dtyp.loc[idx1,'COLUMN_NAME']} data type from VARCHAR2({df1_dtyp.loc[idx1,'CHAR_LENGTH']} CHAR) to VARCHAR2({df2_dtyp.loc[idx2,'CHAR_LENGTH']} CHAR).''')
            cursor.execute(text(f'''ALTER TABLE {schema}.{table_name} MODIFY {df1_dtyp.loc[idx1,'COLUMN_NAME']} VARCHAR2({df2_dtyp.loc[idx2,'CHAR_LENGTH']} CHAR)'''))

def sql_where_condition(condition):
    if condition == None:
        return 'where 1=1'
    elif isinstance(condition,list) == True:
        return 'where ' + ' and '.join(condition)
    else:
        return 'where ' + condition
            
def insert_row_to_sql(df,df_base,schema,table_name,source_owner,source_table):
    import time
    time_elapsed = time.time()
    my_credentials = get_credentials('my_credentials.txt',schema)
    engine = connect_sqlalchemy(my_credentials)
    conn = engine.connect()
    col_intersect = df.columns.intersection(df_base.columns)
    df = df_fix_dtype(df[col_intersect],schema)
    for cols in df.columns:
        if ((check_dtime(df[cols])['is_date'] == True or check_dtime(df[cols])['is_timestamp'] == True) and
            (df[df[cols].notnull()][cols].apply(lambda x: isinstance(x,int)).all() == False)
           ):
            if check_dtime(df[cols])['is_date'] == True and check_dtime(df[cols])['is_timestamp'] == False:
                dt = df[cols].apply(lambda x: f"TO_DATE('{str(x).split(' ')[0]}', 'YYYY-MM-DD')" if is_nan(x) == False else x)
                df[cols] = dt
            elif check_dtime(df[cols])['is_date'] == True and check_dtime(df[cols])['is_timestamp'] == True:
                dt = df[cols].apply(lambda x: f"TO_DATE('{str(x)}', 'YYYY-MM-DD HH24:MI:SS')" if is_nan(x) == False else x)
                df[cols] = dt
    
    # Fix Oracle data types of new and current table before inserting rows.
    fix_ora_col_length(df_base,df,engine,conn,schema,table_name,source_owner,source_table)
    for i in range(df.shape[0]):
        val = df.iloc[i]
        col = df.columns

        _dict = {}
        for idx in range(df.shape[1]):
            if is_nan(val[idx]) == False:
                if isinstance(val[idx],str) == True and 'TO_DATE' not in val[idx]:
                    val[idx] = fix_ora_string(val[idx])
                    val[idx] = fix_ora_string(val[idx],'"','"temp_quotation"')
                _dict.update({col[idx]: val[idx]})
        col = str(list(_dict.keys()))
        col = col.replace("[","")
        col = col.replace("]","")
        col = col.replace("'","")
        val = str(list(_dict.values()))
        val = val.replace("[","")
        val = val.replace("]","")
        val = val.replace('"',"'")
        val = val.replace("'temp_quotation'",'"')
        query = f'INSERT INTO {schema}.{table_name} \n(\n{col}\n)\nVALUES\n(\n{val}\n)'
        query = query.replace('"TO_DATE','TO_DATE')
        query = query.replace("'TO_DATE",'TO_DATE')
        query = query.replace(')",','),')
        query = query.replace("')',","'),")
        try:
            conn.execute(text(query))
        except Exception as err:
            print(f'Row {i} failed.')
            conn.rollback()
            print('Rollback.')
            conn.close()
            engine.dispose()
            time_elapsed = get_time_elapsed(time_elapsed)
            raise
    conn.commit()
    conn.close()
    engine.dispose()
    print('Commit.')
    time_elapsed = get_time_elapsed(time_elapsed)

    
def bulk_insert_row_to_sql(df,df_base,schema,table_name):
    import time
    start_time = time.time()
    time_elapsed = time.time()
    my_credentials = get_credentials('my_credentials.txt',schema)
    engine = connect_sqlalchemy(my_credentials)
    Session = sessionmaker(bind=engine)
    session = Session()
    col_intersect = df.columns.intersection(df_base.columns)
    dtyp = get_sqlalchemy_dtyp_dict(df[col_intersect],engine)
    df = df_fix_dtype(df[col_intersect],schema,dtyp)
    df = excel_df_fix_strftime(df)
    try:
        # Convert dataframe to sql
        event.listens_for(engine, 'before_cursor_execute')
        sql = df.to_sql(name=table_name, con=engine, schema=schema, if_exists='append', dtype=dtyp, index=False)
    except Exception as err:
        session.rollback()
        engine.dispose()
        print('Rollback.')
        raise
    else:
        session.commit()
        print('Commit.')
    end_time = time.time()
    print('Execution time: {:2f}'.format(end_time - start_time),'s')
    
def merge_to_sql(query,query_sample,schema,table_name,connection_key):
    import time
    time_elapsed = time.time()
    my_credentials = get_credentials('my_credentials.txt',schema)
    engine = connect_sqlalchemy(my_credentials)
    conn = engine.connect()
    df = query_to_df(query_sample,schema)
    merge_statement = query_merge(df,query,schema,table_name,connection_key)
    try:
        query = f'''BEGIN
        {merge_statement}
        COMMIT;
EXCEPTION
    WHEN OTHERS THEN
        ROLLBACK;
        RAISE;
END;'''
        conn.execute(text(query))
    except Exception as err:
        conn.rollback()
        conn.close()
        engine.dispose()
        print('Merge failed. Rollback.')
        time_elapsed = get_time_elapsed(time_elapsed)
        raise
    else:
        conn.commit()
        conn.close()
        engine.dispose()
        print('Merged successfully. Commit.')
        time_elapsed = get_time_elapsed(time_elapsed)

def var_merge_to_sql(schema,source_table,source_owner,
                     query_c_date,query_sample,date_start,date_end,
                     connection_key,time_out=60,temporal='days',increment=1,direction=1):
    import time
    from datetime import date
    from dateutil.relativedelta import relativedelta
    # Open
    my_credentials = get_credentials('my_credentials.txt',schema)
    engine = connect_sqlalchemy(my_credentials)
    conn = engine.raw_connection()
    cursor = conn.cursor()
    df = query_to_df(query_sample,schema)
    
    d1 = date_start
    d1_split = d1.split("'")
    if d1_split[-1] == '':
        d1_split.pop(-1)
    d2 = date_end
    d2_split = d2.split("'")
    if d2_split[-1] == '':
        d2_split.pop(-1)
    if "DATE '" in d1 and '-' in d1 and len(d1_split) > 2:
        d1 = "'".join(d1_split[:2])+"'"
        d1_inc = d1_split[-1].strip().split(' ')
        inc = []
        for i in d1_inc:
            if len(d1_inc) == 1:
                d1_inc = int(i)
            else:
                if i == "-":
                    inc.append(i)
                elif i == "+":
                    pass
                elif i == "":
                    pass
                elif type(int(i)) == int:
                    inc.append(i)
        d1_inc = int(''.join(inc))
        d1 = ora_date_to_pydate(d1)
        d1 += relativedelta(days=d1_inc)
        date_start = 'DATE ' + "'" + str(d1) + "'"
        
    if "DATE '" in d2 and '-' in d2 and len(d2_split) > 2:
        d2 = "'".join(d2_split[:2])+"'"
        d2_inc = d2_split[-1].strip().split(' ')
        inc = []
        for i in d2_inc:
            if len(d2_inc) == 1:
                d2_inc = int(i)
            else:
                if i == "-":
                    inc.append(i)
                elif i == "+":
                    pass
                elif i == "":
                    pass
                elif type(int(i)) == int:
                    inc.append(i)
        d2_inc = int(''.join(inc))
        d2 = ora_date_to_pydate(d2)
        d2 += relativedelta(days=d2_inc)
        date_end = 'DATE ' + "'" + str(d2) + "'"
        
    date_today = date.today()
    if ora_date_to_pydate(date_end) > date_today:
        date_end = 'DATE ' + "'" + str(date_today) + "'"
    c_date_start = date_start
    c_date_end = date_end
    date_diff = days_between(c_date_start,c_date_end,temporal)
    
    try:
        print('Timeline: {} to {}'.format(c_date_start,c_date_end))
        time_elapsed = time.time()
        # Set initial value for merge boolean
        is_merge_done = False
        def func_merge(df,source_table,source_owner,cursor,
                       query_c_date,date_start,date_end,connection_key):
            time_elapsed = time.time()
            query = var_query_fstring(query_c_date,date_start,date_end)
            merge_statement = query_merge(df,query,source_owner,source_table,connection_key)
            try:
                query = f'''BEGIN
        {merge_statement}
        COMMIT;
EXCEPTION
    WHEN OTHERS THEN
        ROLLBACK;
        RAISE;
END;'''
                cursor.execute(query)
            except sqlalchemy.exc.DatabaseError as db_err:
                conn.rollback()
                print('Merge failed. Rollback.')
                cursor.close()
                conn.close()
                engine.dispose()
                print('Closed connection.')
                time_elapsed = get_time_elapsed(time_elapsed)
            except KeyboardInterrupt:
                cursor.close()
                conn.close()
                engine.dispose()
                raise
            except Exception as err:
                print('Error: ' + str(err))
                time_elapsed = get_time_elapsed(time_elapsed)
            else:
                conn.commit()
                print('Merged successfully. Commit.')
                time_elapsed = get_time_elapsed(time_elapsed)
        
        from func_timeout import func_timeout, FunctionTimedOut
        try:
            func_timeout(time_out, func_merge, args=(df,source_table,source_owner,cursor,
                                                     query_c_date,date_start,date_end,connection_key))
        except FunctionTimedOut:
            if temporal.upper() == 'DAYS' or temporal.upper() == 'DAY' or temporal.upper() == 'DAILY' or temporal.upper() == 'DD':
                if increment == 1:
                    print(f'Process takes too long. Partitioning timeline by day and populating individually.')
                else:
                    print(f'Process takes too long. Partitioning timeline by {increment} days and populating individually.')
            elif temporal.upper() == 'WEEKS' or temporal.upper() == 'WEEK' or temporal.upper() == 'WEEKLY' or temporal.upper() == 'WW' or temporal.upper() == 'IW':
                if increment == 1:
                    print(f'Process takes too long. Partitioning timeline by week and populating individually.')
                else:
                    print(f'Process takes too long. Partitioning timeline by {increment} weeks and populating individually.')
            elif temporal.upper() == 'MONTHS' or temporal.upper() == 'MONTH' or temporal.upper() == 'MONTHLY' or temporal.upper() == 'MM':
                if increment == 1:
                    print(f'Process takes too long. Partitioning timeline by month and populating individually.')
                else:
                    print(f'Process takes too long. Partitioning timeline by {increment} months and populating individually.')
            elif temporal.upper() == 'YEARS' or temporal.upper() == 'YEAR' or temporal.upper() == 'YEARLY' or temporal.upper() == 'YY'  or temporal.upper() == 'YYYY':
                if increment == 1:
                    print(f'Process takes too long. Partitioning timeline by year and populating individually.')
                else:
                    print(f'Process takes too long. Partitioning timeline by {increment} years and populating individually.')
            time_elapsed = get_time_elapsed(time_elapsed)
        except Exception as err:
            print('Error: ' + str(err))
            time_elapsed = get_time_elapsed(time_elapsed)
        else:
            is_merge_done = True
        
        chunk_merge_start_time = time.time()
        if is_merge_done == False:
            if direction == -1:
                c_date_start = ora_date_inc(c_date_end,temporal,direction*increment)
            else:
                c_date_end = ora_date_inc(c_date_start,temporal,increment)
            is_sample_done = False
            for i in range(date_diff):
                try:
                    if is_sample_done == False:
                        if direction == -1:
                            c_date_start = ora_date_inc(c_date_end,'day',direction*1)
                        else:
                            c_date_end = ora_date_inc(c_date_start,'day',1)
                        print('Timeline: {} to {}'.format(c_date_start,c_date_end))
                        func_merge(df,source_table,source_owner,cursor,
                                   query_c_date,c_date_start,c_date_end,connection_key)
                        if direction == -1:
                            if c_date_start != date_start:
                                c_date_end = c_date_start
                        else:
                            if c_date_end != date_end:
                                c_date_start = c_date_end
                        is_sample_done = True
                    
                    if direction == -1:
                        c_date_start = ora_date_inc(c_date_end,temporal,direction*increment)
                    else:
                        c_date_end = ora_date_inc(c_date_start,temporal,increment)
                    if ora_date_to_pydate(c_date_end) > ora_date_to_pydate(date_end):
                        if direction == -1:
                            c_date_start = date_start
                            if c_date_start == c_date_end:
                                c_date_start = ora_date_inc(c_date_start,'days',direction*1)
                        else:
                            c_date_end = date_end
                            if c_date_end == c_date_start:
                                c_date_end = ora_date_inc(c_date_end,'days',1)
                    print('Timeline: {} to {}'.format(c_date_start,c_date_end))
                    func_merge(df,source_table,source_owner,cursor,
                               query_c_date,c_date_start,c_date_end,connection_key)
                except Exception as err:
                    break
                    raise
                else:
                    if direction == -1:
                        c_date_end = c_date_start
                        c_date_start = ora_date_inc(c_date_start,temporal,direction*increment)
                        if ora_date_to_pydate(c_date_start) < ora_date_to_pydate(date_start):
                            c_date_start = date_start
                            if c_date_start == c_date_end:
                                c_date_start = ora_date_inc(c_date_start,'days',direction*1)
                    else:
                        c_date_start = c_date_end
                        c_date_end = ora_date_inc(c_date_end,temporal,increment)
                        if ora_date_to_pydate(c_date_end) > ora_date_to_pydate(date_end):
                            c_date_end = date_end
                            if c_date_end == c_date_start:
                                c_date_end = ora_date_inc(c_date_end,'days',1)
        chunk_merge_end_time = time.time()
        print('Chunks merge done. Time elapsed: {:2f}'.format(chunk_merge_end_time - chunk_merge_start_time),'s')
        cursor.close()
        conn.close()
        engine.dispose()
    except KeyboardInterrupt:
        print('Keyboard Interrupted.')
    except Exception as err:
        conn.rollback()
        conn.close()
        engine.dispose()
        
def split_dataframe(df, chunk_size): 
    chunks = list()
    num_chunks = len(df) // chunk_size + 1
    for i in range(num_chunks):
        chunks.append(df[i*chunk_size:(i+1)*chunk_size])
    return chunks

def query_to_excel(schema,query,fname,sheet_name=None,chunk_size=1000000):
    my_credentials = get_credentials('my_credentials.txt',schema)
    fname_check_result = fname_check(fname,schema)
    is_file = fname_check_result['is_file']
    fname = fname_check_result['fname']
    folder_path = fname_check_result['folder_path']
    table_name = fname_check_result['table_name']
    
    if sheet_name == None:
        sheet_name = 'Sheet1'
        
    start_time = time.time()
    
    if is_file == True:
        print('File already exists.')
        is_overwrite = input('Do you want to overwrite the existing file? [YES/NO]  |  ')
        if is_overwrite.upper() == 'YES' or is_overwrite.upper() == 'Y':
            # Open
            engine = connect_sqlalchemy(my_credentials)
            try:
                start_time = time.time()
                if chunk_size != None:
                    # Create empty list
                    dfl = []
                    # Create empty dataframe
                    dfs = pd.DataFrame()
                    # Start Chunking SQL
                    chunk_idx = 0
                    start_time_file = time.time()
                    start_time_chunk = time.time()
                    for chunk in pd.read_sql_query(query,engine,chunksize=chunk_size):
                        # Start Appending Data Chunks from SQL Result set into List
                        dfl.append(chunk)
                        chunk_idx += 1
                        end_time_chunk = time.time()
                        print('CHUNK {} appended to list. Time elapsed: {:2f}'.format(chunk_idx,end_time_chunk - start_time_chunk))
                        start_time_chunk = time.time()
                    end_time_file = time.time()
                    engine.dispose()
                    print('Chunking done. Time elapsed: {:2f}'.format(end_time_file - start_time_file),'s')
                    # Append data from list to dataframe
                    start_time_df = time.time()
                    dfs = pd.concat(dfl, ignore_index=True)
                    dfs.columns = dfs.columns.str.upper()
                    dfs.where(pd.notnull(dfs),np.nan,inplace=True)
                    dfs = df_fix_dtype(dfs,schema)
                    del dfl # delete variable to clear memory cache
                    end_time_df = time.time()
                    print('Dataframe created. Time elapsed: {:2f}'.format(end_time_df - start_time_df),'s')
                else:
                    start_time_df = time.time()
                    dfs = pd.read_sql_query(query,engine)
                    dfs.columns = dfs.columns.str.upper()
                    dfs.where(pd.notnull(dfs),np.nan,inplace=True)
                    dfs = df_fix_dtype(dfs,schema)
                    end_time_df = time.time()
                    print('Dataframe created. Time elapsed: {:2f}'.format(end_time_df - start_time_df),'s')

                wb = Workbook()
                # Chunk dataframe and insert into spreah sheet separately
                if dfs.shape[0] >= chunk_size:
                    start_time_file = time.time()
                    wb = Workbook()
                    chunks = split_dataframe(dfs,chunk_size)
                    for i in range(len(chunks)):
                        start_time_chunk = time.time()
                        df_col = [chunks[i].columns]
                        df_data = chunks[i].values.tolist()
                        df_data = [['' if (str(val).upper() == 'NAN') or (str(val).upper() == 'NONE') or (str(val).upper() == 'NAT') else val for val in row] for row in df_data]
                        values = df_col + df_data
                        ws = wb.new_sheet('CHUNK_'+str(i+1), data=values)
                        chunks[i] = None # replace chunk with None to clear memory cache
                        end_time_chunk = time.time()
                        print('New sheet created for CHUNK {}. Time elapsed: {:2f}'.format(i+1,end_time_chunk - start_time_chunk))
                    del chunks # delete variable to clear memory cache
                    wb.save(fname)
                    end_time_file = time.time()
                    print('Workbook created. Time elapsed: {:2f}'.format(end_time_file - start_time_file),'s')
                else:
                    start_time_file = time.time()
                    wb = Workbook()
                    df_col = [dfs.columns]
                    df_data = dfs.values.tolist()
                    df_data = [['' if (str(val).upper() == 'NAN') or (str(val).upper() == 'NONE') or (str(val).upper() == 'NAT') else val for val in row] for row in df_data]
                    values = df_col + df_data
                    ws = wb.new_sheet(sheet_name, data=values)
                    wb.save(fname)
                    end_time_file = time.time()
                    print('Workbook created. Time elapsed: {:2f}'.format(end_time_file - start_time_file),'s')
                end_time = time.time()
                print('Execution time: {:2f}'.format(end_time - start_time),'s')
            except Exception as err:
                print('Error: ' + str(err))
                engine.dispose()
    else:
        # Open
        engine = connect_sqlalchemy(my_credentials)
        try:
            start_time = time.time()
            if chunk_size != None:
                # Create empty list
                dfl = []
                # Create empty dataframe
                dfs = pd.DataFrame()
                # Start Chunking SQL
                chunk_idx = 0
                start_time_file = time.time()
                start_time_chunk = time.time()
                for chunk in pd.read_sql_query(query,engine,chunksize=chunk_size):
                    # Start Appending Data Chunks from SQL Result set into List
                    dfl.append(chunk)
                    chunk_idx += 1
                    end_time_chunk = time.time()
                    print('CHUNK {} appended to list. Time elapsed: {:2f}'.format(chunk_idx,end_time_chunk - start_time_chunk))
                    start_time_chunk = time.time()
                end_time_file = time.time()
                engine.dispose()
                print('Chunking done. Time elapsed: {:2f}'.format(end_time_file - start_time_file),'s')
                # Append data from list to dataframe
                start_time_df = time.time()
                dfs = pd.concat(dfl, ignore_index=True)
                dfs.columns = dfs.columns.str.upper()
                dfs.where(pd.notnull(dfs),np.nan,inplace=True)
                dfs = df_fix_dtype(dfs,schema)
                end_time_df = time.time()
                print('Dataframe created. Time elapsed: {:2f}'.format(end_time_df - start_time_df),'s')
            else:
                start_time_df = time.time()
                dfs = pd.read_sql_query(query,engine)
                dfs.columns = dfs.columns.str.upper()
                dfs.where(pd.notnull(dfs),np.nan,inplace=True)
                dfs = df_fix_dtype(dfs,schema)
                end_time_df = time.time()
                print('Dataframe created. Time elapsed: {:2f}'.format(end_time_df - start_time_df),'s')
                

            wb = Workbook()
            # Chunk dataframe and insert into spreah sheet separately
            if dfs.shape[0] >= chunk_size:
                start_time_file = time.time()
                wb = Workbook()
                chunks = split_dataframe(dfs,chunk_size)
                for i in range(len(chunks)):
                    start_time_chunk = time.time()
                    df_col = [chunks[i].columns]
                    df_data = chunks[i].values.tolist()
                    df_data = [['' if (str(val).upper() == 'NAN') or (str(val).upper() == 'NONE') or (str(val).upper() == 'NAT') else val for val in row] for row in df_data]
                    values = df_col + df_data
                    ws = wb.new_sheet('CHUNK_'+str(i+1), data=values)
                    end_time_chunk = time.time()
                    print('New sheet created for CHUNK {}. Time elapsed: {:2f}'.format(i+1,end_time_chunk - start_time_chunk))
                del chunks # delete variable to clear memory cache
                fname = folder_path+'/'+table_name+''.join(pathlib.Path(fname).suffixes)
                wb.save(fname)
                end_time_file = time.time()
                print('Workbook created. Time elapsed: {:2f}'.format(end_time_file - start_time_file),'s')
            else:
                start_time_file = time.time()
                wb = Workbook()
                df_col = [dfs.columns]
                df_data = dfs.values.tolist()
                df_data = [['' if (str(val).upper() == 'NAN') or (str(val).upper() == 'NONE') or (str(val).upper() == 'NAT') else val for val in row] for row in df_data]
                values = df_col + df_data
                ws = wb.new_sheet(sheet_name, data=values)
                fname = folder_path+'/'+table_name+''.join(pathlib.Path(fname).suffixes)
                wb.save(fname)
                end_time_file = time.time()
                print('Workbook created. Time elapsed: {:2f}'.format(end_time_file - start_time_file),'s')
            end_time = time.time()
            print('Execution time: {:2f}'.format(end_time - start_time),'s')
        except Exception as err:
            print('Error: ' + str(err))
            engine.dispose()
    return dfs

def df_to_excel(df,schema,fname,sheet_name=None,chunk_size=1000000,dtyp=None):
    my_credentials = get_credentials('my_credentials.txt',schema)
    fname_check_result = fname_check(fname,schema)
    is_file = fname_check_result['is_file']
    fname = fname_check_result['fname']
    folder_path = fname_check_result['folder_path']
    table_name = fname_check_result['table_name']
    
    if sheet_name == None:
        sheet_name = 'Sheet1'
    
    df.where(pd.notnull(df),np.nan,inplace=True)
    df = df_fix_dtype(df,schema,dtyp=dtyp)
    start_time = time.time()
    
    if is_file == True:
        print('File already exists.')
        is_overwrite = input('Do you want to overwrite the existing file? [YES/NO]  |  ')
        if is_overwrite.upper() == 'YES' or is_overwrite.upper() == 'Y':
            try:                    
                wb = Workbook()
                # Chunk dataframe and insert into spreah sheet separately
                if df.shape[0] >= chunk_size:
                    start_time_file = time.time()
                    wb = Workbook()
                    chunks = split_dataframe(df,chunk_size)
                    for i in range(len(chunks)):
                        start_time_chunk = time.time()
                        df_col = [chunks[i].columns]
                        df_data = chunks[i].values.tolist()
                        df_data = [['' if (str(val).upper() == 'NAN') or (str(val).upper() == 'NONE') or (str(val).upper() == 'NAT') else val for val in row] for row in df_data]
                        values = df_col + df_data
                        ws = wb.new_sheet('CHUNK_'+str(i+1), data=values)
                        end_time_chunk = time.time()
                        print('New sheet created for CHUNK {}. Time elapsed: {:2f}'.format(i+1,end_time_chunk - start_time_chunk))
                    del chunks # delete variable to clear memory cache
                    wb.save(fname)
                    end_time_file = time.time()
                    print('Workbook created. Time elapsed: {:2f}'.format(end_time_file - start_time_file),'s')
                else:
                    start_time_file = time.time()
                    wb = Workbook()
                    df_col = [df.columns]
                    df_data = df.values.tolist()
                    df_data = [['' if (str(val).upper() == 'NAN') or (str(val).upper() == 'NONE') or (str(val).upper() == 'NAT') else val for val in row] for row in df_data]
                    values = df_col + df_data
                    ws = wb.new_sheet(sheet_name, data=values)
                    wb.save(fname)
                    end_time_file = time.time()
                    print('Workbook created. Time elapsed: {:2f}'.format(end_time_file - start_time_file),'s')
            except Exception as err:
                print('Error: ' + str(err))
        elif is_overwrite.upper() == 'NO' or is_overwrite.upper() == 'N':
            try:
                start_time_file = time.time()
                # Load the existing workbook
                book = load_workbook(fname)
                # Open an ExcelWriter object and specify the engine to use openpyxl
                with pd.ExcelWriter(folder_path, engine='openpyxl') as writer:
                    writer.book = book
                    # Add the new DataFrame as a new sheet
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    # Save the workbook
                    writer.save()
                    end_time_file = time.time()
                    print('Sheet appended. Time elapsed: {:2f}'.format(end_time_file - start_time_file),'s')
            except Exception as err:
                print('Error: ' + str(err))
    else:
        try:                    
            wb = Workbook()
            # Chunk dataframe and insert into spreah sheet separately
            if df.shape[0] >= chunk_size:
                start_time_file = time.time()
                wb = Workbook()
                chunks = split_dataframe(df,chunk_size)
                for i in range(len(chunks)):
                    start_time_chunk = time.time()
                    df_col = [chunks[i].columns]
                    df_data = chunks[i].values.tolist()
                    df_data = [['' if (str(val).upper() == 'NAN') or (str(val).upper() == 'NONE') or (str(val).upper() == 'NAT') else val for val in row] for row in df_data]
                    values = df_col + df_data
                    ws = wb.new_sheet('CHUNK_'+str(i+1), data=values)
                    end_time_chunk = time.time()
                    print('New sheet created for CHUNK {}. Time elapsed: {:2f}'.format(i+1,end_time_chunk - start_time_chunk))
                del chunks # delete variable to clear memory cache
                fname = folder_path+'/'+table_name+''.join(pathlib.Path(fname).suffixes)
                wb.save(fname)
                end_time_file = time.time()
                print('Workbook created. Time elapsed: {:2f}'.format(end_time_file - start_time_file),'s')
            else:
                start_time_file = time.time()
                wb = Workbook()
                df_col = [df.columns]
                df_data = df.values.tolist()
                df_data = [['' if (str(val).upper() == 'NAN') or (str(val).upper() == 'NONE') or (str(val).upper() == 'NAT') else val for val in row] for row in df_data]
                values = df_col + df_data
                ws = wb.new_sheet(sheet_name, data=values)
                fname = folder_path+'/'+table_name+''.join(pathlib.Path(fname).suffixes)
                wb.save(fname)
                end_time_file = time.time()
                print('Workbook created. Time elapsed: {:2f}'.format(end_time_file - start_time_file),'s')
        except Exception as err:
            print('Error: ' + str(err))

def get_py_dtyp_as_dict(df):
    import datetime
    from datetime import time
    from datetime import date
    
    df_dtyp = pd.DataFrame(columns=['COLUMN_NAME','DTYP'])
    df_dtyp['COLUMN_NAME'] = df.columns
    for cols,rows in df.loc[:10000,:].items():
        idx = df_dtyp.index[df_dtyp['COLUMN_NAME']==cols]
        is_int_type = rows[rows.notnull()].apply(lambda x: isinstance(x,int)).any()
        is_int64_type = rows[rows.notnull()].apply(lambda x: isinstance(x,np.int64)).any()
        is_float_type = rows[rows.notnull()].apply(lambda x: isinstance(x,float)).any()
        is_float64_type = rows[rows.notnull()].apply(lambda x: isinstance(x,np.float64)).any()
        is_date_type = check_dtime(rows)['is_date']
        is_datetime_type = check_dtime(rows)['is_timestamp']
        is_str_type = rows[rows.notnull()].apply(lambda x: isinstance(x,str)).any()
        unique_dtyp = rows[rows.notnull()].apply(lambda x: type(x)).unique()


        if is_int_type == True or is_int64_type == True:
            df_dtyp.loc[idx,'DTYP'] = int
        elif is_float_type == True or is_float64_type == True:
            df_dtyp.loc[idx,'DTYP'] = float
        elif is_datetime_type == True:
            df_dtyp.loc[idx,'DTYP'] = datetime
        elif is_date_type == True:
            df_dtyp.loc[idx,'DTYP'] = date
        elif 'dtime' in str(cols).lower():
            df_dtyp.loc[idx,'DTYP'] = datetime
        elif 'date' in str(cols).lower():
            df_dtyp.loc[idx,'DTYP'] = date
        elif is_str_type == True:
            df_dtyp.loc[idx,'DTYP'] = str
        else:
            if len(unique_dtyp) == 0:
                df_dtyp.loc[idx,'DTYP'] = float
            else:
                df_dtyp.loc[idx,'DTYP'] = unique_dtyp[0]
                
    return dict(df_dtyp.values)
            
def get_sqlalchemy_dtyp_dict(df,engine,source_owner='TEST_OWNER',source_table='TEST_TABLE'):
    import time
    from datetime import date
    from datetime import datetime
    from sqlalchemy import types
    from sqlalchemy.dialects.oracle import (
                                            DATE,
                                            TIMESTAMP,
                                            NUMBER,
                                            VARCHAR2
                                            )
    time_elapsed = time.time()
    df_sample = df.loc[:10000,:]
    df = df_sample[df_sample.notnull()]
    df_col = df.columns
    df_col_typ = df.dtypes
    df_col_list = [cols for cols in df_col]
    df_col_str = str(df_col_list)
    df_col_str = df_col_str.replace("[","")
    df_col_str = df_col_str.replace("]","")
    df_col_str = df_col_str.replace(", ",", \n")
    dtyp = []
    print('Fetching Oracle data types.')
    df_col_oracle_typ = get_oracle_data_type(df_col_str,engine,source_owner=source_owner,source_table=source_table)
    time_elapsed = get_time_elapsed(time_elapsed)
    print('Fetching python data types.')
    py_dtyp_dict = get_py_dtyp_as_dict(df)
    time_elapsed = get_time_elapsed(time_elapsed)
    print('Converting to sqlalchemy data types.')
    for cols in df_col:
        if cols in df_col_oracle_typ['COLUMN_NAME'].values:
            get_index = df_col_oracle_typ.index[df_col_oracle_typ['COLUMN_NAME']==cols]
            oracle_dtyp = str(df_col_oracle_typ['DATA_TYPE'][get_index].values)
            oracle_dtyp = oracle_dtyp.replace("[","")
            oracle_dtyp = oracle_dtyp.replace("]","")
            pair = []
            pair.append(cols)
            pair.append(oracle_dtyp)
            dtyp.append(pair)
        elif py_dtyp_dict[cols] == datetime:
            pair = []
            pair.append(cols)
            pair.append('TIMESTAMP')
            dtyp.append(pair)
        elif py_dtyp_dict[cols] == date:
            pair = []
            pair.append(cols)
            pair.append('DATE')
            dtyp.append(pair)
        elif py_dtyp_dict[cols] == int:
            pair = []
            pair.append(cols)
            pair.append('INTEGER')
            dtyp.append(pair)
        elif py_dtyp_dict[cols] == float:
            pair = []
            pair.append(cols)
            pair.append('FLOAT')
            dtyp.append(pair)
        elif py_dtyp_dict[cols] == str:
            pair = []
            pair.append(cols)
            pair.append('VARCHAR')
            dtyp.append(pair)
        else:
            pair = []
            pair.append(cols)
            pair.append('VARCHAR2(80 CHAR)')
            dtyp.append(pair)
    df_col_len = pd.DataFrame(columns=['COLUMN_NAME','CHAR_LENGTH'])
    df_col_len['COLUMN_NAME'] = df_col
    for idx in range(df_col_len.shape[0]):
        df_col_len.loc[idx,'CHAR_LENGTH'] = df[df_col_len.loc[idx,'COLUMN_NAME']].astype(str).map(len).max()
    df_col_len.set_index('COLUMN_NAME', inplace=True)
    dtyp = dict(dtyp)
    data = {p:[q] for p,q in dtyp.items()}
    df_dict = pd.DataFrame(data).transpose()
    df_dict.rename(columns={0:'DTYP'}, inplace=True)
    df_dtyp = pd.DataFrame(columns=['COLUMN_NAME','DTYP'])
    df_dtyp['COLUMN_NAME'] = df_col
    for idx in range(df_dtyp.shape[0]):
        col_dtyp = df_dict.loc[df_dtyp.loc[idx,'COLUMN_NAME']].values[0]
        if 'TIMESTAMP' in col_dtyp:
            df_dtyp.loc[idx,'DTYP'] = TIMESTAMP(6)
        elif 'DATE' in col_dtyp:
            df_dtyp.loc[idx,'DTYP'] = DATE
        elif 'NUMBER' in col_dtyp or 'INTEGER' in col_dtyp or 'FLOAT' in col_dtyp:
            df_dtyp.loc[idx,'DTYP'] = NUMBER
        else:
            varchar_len = df_col_len.loc[df_dtyp.loc[idx,'COLUMN_NAME']].values[0]
            df_dtyp.loc[idx,'DTYP'] = VARCHAR2(varchar_len)
    return dict(df_dtyp.values)
    
def get_ora_dtyp_dict(df,engine,source_owner='TEST_OWNER',source_table='TEST_TABLE'):
    df_col = df.columns
    df_col_typ = df.dtypes
    df_col_list = [df_col[i] for i in range(df.shape[1])]
    df_col_str = str(df_col_list)
    df_col_str = df_col_str.replace("[","")
    df_col_str = df_col_str.replace("]","")
    df_col_str = df_col_str.replace(", ",", \n")
    dtyp = []
    df_col_oracle_typ = get_oracle_data_type(df_col_str,engine,source_owner=source_owner,source_table=source_table)
    for i in range(len(df_col)):
        if 'date' in str(df_col_typ[i]).lower() or 'dtime' in str(df_col[i]).lower():
            pair = []
            pair.append(df_col[i])
            pair.append('DATE')
            dtyp.append(pair)
        elif df_col[i] in df_col_oracle_typ['COLUMN_NAME'].values:
            get_index = df_col_oracle_typ.index[df_col_oracle_typ['COLUMN_NAME']==df_col[i]]
            oracle_dtyp = str(df_col_oracle_typ['DATA_TYPE'][get_index].values)
            oracle_dtyp = oracle_dtyp.replace("[","")
            oracle_dtyp = oracle_dtyp.replace("]","")
            pair = []
            pair.append(df_col[i])
            pair.append(oracle_dtyp)
            dtyp.append(pair)
        elif 'int' in str(df_col_typ[i]) or 'float' in str(df_col_typ[i]) or 'num' in str(df_col_typ[i]):
            pair = []
            pair.append(df_col[i])
            pair.append('NUMBER')
            dtyp.append(pair)
        else:
            pair = []
            pair.append(df_col[i])
            pair.append('VARCHAR2(80 CHAR)')
            dtyp.append(pair)
    dtyp = dict(dtyp)
    return dtyp

def get_ora_dtyp_varchar_dict(df,engine,source_owner='TEST_OWNER',source_table='TEST_TABLE'):
    df_col = df.columns
    df_col_typ = df.dtypes
    df_col_list = [df_col[i] for i in range(df.shape[1])]
    df_col_str = str(df_col_list)
    df_col_str = df_col_str.replace("[","")
    df_col_str = df_col_str.replace("]","")
    df_col_str = df_col_str.replace(", ",", \n")
    dtyp = []
    df_col_oracle_typ = get_oracle_data_type(df_col_str,engine,source_owner=source_owner,source_table=source_table)
    for i in range(len(df_col)):
        if 'date' in str(df_col_typ[i]).lower() or 'dtime' in str(df_col[i]).lower():
            pair = []
            pair.append(df_col[i])
            pair.append('DATE')
            dtyp.append(pair)
        elif df_col[i] in df_col_oracle_typ['COLUMN_NAME'].values:
            get_index = df_col_oracle_typ.index[df_col_oracle_typ['COLUMN_NAME']==df_col[i]]
            oracle_dtyp = str(df_col_oracle_typ['DATA_TYPE'][get_index].values)
            oracle_dtyp = oracle_dtyp.replace("[","")
            oracle_dtyp = oracle_dtyp.replace("]","")
            pair = []
            pair.append(df_col[i])
            pair.append(oracle_dtyp)
            dtyp.append(pair)
        elif 'int' in str(df_col_typ[i]) or 'float' in str(df_col_typ[i]) or 'num' in str(df_col_typ[i]):
            pair = []
            pair.append(df_col[i])
            pair.append('NUMBER')
            dtyp.append(pair)
        else:
            pair = []
            pair.append(df_col[i])
            pair.append('VARCHAR2(80 CHAR)')
            dtyp.append(pair)
    df_col_len = pd.DataFrame(columns=['COLUMN_NAME','CHAR_LENGTH'])
    df_col_len['COLUMN_NAME'] = df_col
    for i in range(df_col_len.shape[0]):
        df_col_len['CHAR_LENGTH'][i] = df[df_col_len['COLUMN_NAME'][i]].astype(str).map(len).max()
    df_col_len.set_index('COLUMN_NAME', inplace=True)
    for i in range(len(dtyp)):
        if 'VARCHAR2(' in dtyp[i][1]:
            dtyp_str = str(dtyp[i][0]).replace("/","")
            dtyp_str = dtyp_str.replace("'","")
            dtyp_str = str(df_col_len.loc[dtyp_str].values)
            dtyp_str = dtyp_str.replace("[","")
            dtyp_str = dtyp_str.replace("]","")
            dtyp[i][1] = str('VARCHAR(' + dtyp_str + ' CHAR)')
    dtyp = dict(dtyp)
    return dtyp

def is_date(string):
    from dateutil.parser import parse
    try:
        parse(string)
        return True
    except ValueError:
        return False
    except Exception as err:
        print('Error: ' + str(err))

def is_date_valid(string):
    import datetime
    from datetime import time
    from datetime import datetime

    def is_date_valid_1(string):
        try:
            string = datetime.strptime(string,'%m/%d/%Y')
            return True,string
        except ValueError:
            return False,string
        except TypeError:
            return False,string
        except Exception as err:
            print('Error: ' + str(err))

    def is_date_valid_2(string):
        try:
            string = datetime.strptime(string,'%Y-%m-%d')
            return True,string
        except ValueError:
            return False,string
        except TypeError:
            return False,string
        except Exception as err:
            print('Error: ' + str(err))

    def is_date_valid_3(string):
        try:
            string = datetime.strptime(string,'%Y-%m-%d %H:%M:%S')
            return True,string
        except ValueError:
            return False,string
        except TypeError:
            return False,string
        except Exception as err:
            print('Error: ' + str(err))

    def is_date_valid_final(string):
        try:
            if is_date_valid_1(string)[0] == True:
                return is_date_valid_1(string)[0],is_date_valid_1(string)[1]
            elif is_date_valid_2(string)[0] == True:
                return is_date_valid_2(string)[0],is_date_valid_2(string)[1]
            elif is_date_valid_3(string)[0] == True:
                return is_date_valid_3(string)[0],is_date_valid_3(string)[1]
            else:
                datetime.strptime(string,'%Y-%m-%d %H:%M:%S')
        except ValueError:
            return False,string
        except TypeError:
            return False,string
        except Exception as err:
            print('Error: ' + str(err))
    return is_date_valid_final(string)[0],is_date_valid_final(string)[1]
        
def is_timestamp(date):
    from datetime import time
    try:
        date.time()
    except ValueError:
        return False # NaTType does not support time
    except AttributeError:
        return False # if object has no time attribute (e.g. str,float,int,datetime.datetime)
    else:
        if date.time() == time(0,0,0):
            return False # pandas.timestamps.Timestamp but with 00:00:00 time, hence date
        else:
            return True
    
def is_datetime(date):
    import datetime
    from datetime import time
    try:
        date.date()
    except ValueError:
        return False # NaTType does not support time
    except AttributeError:
        return False # if object has no time attribute (e.g. str,float,int,datetime.datetime)
    else:
        if date.time() == time(0,0,0):
            return True
        else:
            return False

def check_dtime(series):
    from datetime import time
    notnull = series.apply(lambda x: np.nan if is_nan(x) == True else x)
    notnull_to_dt = pd.to_datetime(notnull,format='%Y-%m-%d %H:%M:%S',errors='coerce')
    notnull_to_dt = notnull_to_dt.loc[notnull_to_dt.notnull()]
    is_date = notnull_to_dt.apply(lambda x: True if x.time() == time(0,0,0) else False).all()
    is_timestamp = notnull_to_dt.apply(lambda x: True if x.time() != time(0,0,0) else False).any()
    
    if is_date == True and len(notnull_to_dt) != 0:
        is_date = True
        is_timestamp = False
    elif is_date == True and is_timestamp == True:
        is_date = True
        is_timestamp = True
    elif 'date' in series.name.lower():
        is_date = True
        is_timestamp = False
    elif 'dtime' in series.name.lower():
        is_date = True
        is_timestamp = True
    elif 'time' in series.name.lower() and 'dtime' not in series.name.lower():
        is_date = False
        is_timestamp = True
    else:
        is_date = False
        is_timestamp = False
    
    _dict_ = {
        'is_date': is_date,
        'is_timestamp': is_timestamp
    }
    return _dict_
        
def is_nan(val):
    if str(val) == 'nan':
        return True
    elif str(val) == 'NaN':
        return True
    elif str(val) == 'None':
        return True
    elif str(val) == '':
        return True
    else:
        return False

def excel_df_fix_strftime(df):
    from datetime import date
    df_sample = df.loc[:10000,:]
    df_sample = df_sample[df_sample.notnull()]
    df = df[df.notnull()]
    for cols in df.columns:
        if ((check_dtime(df_sample[cols])['is_date'] == True or check_dtime(df_sample[cols])['is_timestamp'] == True) and
            (df_sample[df_sample[cols].notnull()][cols].apply(lambda x: isinstance(x,int)).all() == False)
           ):
            if check_dtime(df_sample[cols])['is_date'] == True and check_dtime(df_sample[cols])['is_timestamp'] == False:
                dt = pd.to_datetime(df[cols],format="%Y-%m-%d %H:%M:%S",errors='coerce')
                dt = dt.apply(lambda x: x.date())
                df[cols] = dt
            elif check_dtime(df_sample[cols])['is_date'] == True and check_dtime(df_sample[cols])['is_timestamp'] == True:
                dt = pd.to_datetime(df[cols],format="%Y-%m-%d %H:%M:%S",errors='coerce')
                df[cols] = dt
    return df

def excel_to_sql(fname,schema,table_name,sheet_name,dtyp=None):
    import time
    start_time = time.time()
    time_elapsed = start_time
    
    # Set credential
    my_credentials = get_credentials('my_credentials.txt',schema)
    
    fname_check_result = fname_check(fname,schema,table_name)
    is_file = fname_check_result['is_file']
    fname = fname_check_result['fname']
    
    if table_name == '' or table_name == None:
        table_name = fname_check_result[3]
    else:
        table_name = table_name.upper()
                    
    # Open
    engine = connect_sqlalchemy(my_credentials)
    
    if is_file == True:
        # Read file as dataframe
        if fname.endswith('.xlsx') or fname.endswith('.xlsm') or fname.endswith('.xls'):
            print('Reading file as dataframe.')
            df = pd.read_excel(fname, sheet_name=sheet_name, index_col=False, engine='calamine')
            time_elapsed = get_time_elapsed(time_elapsed)
            if sheet_name == None:
                for sheetnames in df.keys():
                    try:
                        df[sheetnames].columns = df[sheetnames].columns.str.upper()
                        if len(df.keys()) > 1:
                            table_name = table_name+'_'+sheetnames
                        print('Pre-processing data types.')
                        if dtyp == None:
                            dtyp = get_sqlalchemy_dtyp_dict(df[sheetnames],engine)
                        else:
                            dtyp = dtyp
                        df[sheetnames] = df_fix_dtype(df[sheetnames],schema,dtyp=dtyp)
                        time_elapsed = get_time_elapsed(time_elapsed)
                        print('Fixing string datetime.')
                        df[sheetnames] = excel_df_fix_strftime(df[sheetnames])
                        time_elapsed = get_time_elapsed(time_elapsed)
                    except Exception as err:
                        print('Error: ' + str(err))
                        engine.dispose()
                        time_elapsed = get_time_elapsed(time_elapsed)
                    else:
                        print('Importing to database.')
                        try:
                            # Convert dataframe to sql
                            event.listens_for(engine, 'before_cursor_execute')
                            sql = df[sheetnames].to_sql(name=table_name, con=engine, schema=schema, dtype=dtyp, index=False, chunksize=10000)
                            end_time = time.time()
                            print('Execution time: {:2f}'.format(end_time - start_time),'s')
                            # Close
                            engine.dispose()
                        except ValueError:
                            engine.dispose()
                            display(dtyp)
                            time_elapsed = get_time_elapsed(time_elapsed)
                            raise
                        except Exception as err:
                            print('Error: ' + str(err))
                            engine.dispose()
                            time_elapsed = get_time_elapsed(time_elapsed)
            else:
                try:
                    df.columns = df.columns.str.upper()
                    print('Pre-processing data types.')
                    try:
                        if dtyp == None:
                            dtyp = get_sqlalchemy_dtyp_dict(df,engine)
                        else:
                            dtyp = dtyp
                        df = df_fix_dtype(df,schema,dtyp=dtyp)
                        time_elapsed = get_time_elapsed(time_elapsed)
                        print('Fixing string datetime.')
                        df = excel_df_fix_strftime(df)
                        time_elapsed = get_time_elapsed(time_elapsed)
                    except:
                        engine.dispose()
                except Exception as err:
                    print('Error: ' + str(err))
                    engine.dispose()
                    time_elapsed = get_time_elapsed(time_elapsed)
                else:
                    print('Importing to database.')
                    try:
                        # Convert dataframe to sql
                        for idx, (key,value) in enumerate(dtyp.items()):
                            if 'VARCHAR' in str(value) and key == df.columns[idx]:
                                df[key] = df[key].apply(lambda x: '' if is_nan(x) == True else str(x))
                        event.listens_for(engine, 'before_cursor_execute')
                        '''
                        col_sample is for testing Database Error DPI 1001: out of memory. When using pd.read_sql_query()
                        and importing the generated dataframe back into database, there are no issues. However, when the
                        generated dataframe is exported into a spreadsheet, re-opened via pandas, and imported back to
                        database, there will be memory error due to large number of columns. Workaround will be to limit 
                        chunksize parameter as small as possible or upgrade to higher RAM.
                        fix it.
                        '''
#                         df_copy = df.copy()
#                         col_sample = []
#                         for idx,col in enumerate(df_copy.columns):
#                             if idx < 10:
#                                 col_sample.append(col)
#                         df_copy[col_sample]
#                         df = df_copy[col_sample]
                        sql = df.to_sql(name=table_name, con=engine, schema=schema, dtype=dtyp, index=False, chunksize=10000)
                        end_time = time.time()
                        print('Execution time: {:2f}'.format(end_time - start_time),'s')
                        # Close
                        engine.dispose()
                    except ValueError as err:
                        engine.dispose()
                        display(dtyp)
                        time_elapsed = get_time_elapsed(time_elapsed)
                    except Exception as err:
                        print('Error: ' + str(err))
                        engine.dispose()
                        time_elapsed = get_time_elapsed(time_elapsed)
        elif fname.endswith('.csv'):
            try:
                print('Reading file as dataframe.')
                df = pd.read_csv(fname, index_col=False)
                time_elapsed = get_time_elapsed(time_elapsed)
                df.columns = df.columns.str.upper()
                print('Pre-processing data types.')
                if dtyp == None:
                    dtyp = get_sqlalchemy_dtyp_dict(df,engine)
                else:
                    dtyp = dtyp
                df = df_fix_dtype(df,schema,dtyp=dtyp)
                time_elapsed = get_time_elapsed(time_elapsed)
                print('Fixing string datetime.')
                df = excel_df_fix_strftime(df)
                time_elapsed = get_time_elapsed(time_elapsed)
            except Exception as err:
                print('Error: ' + str(err))
                engine.dispose()
                time_elapsed = get_time_elapsed(time_elapsed)
            else:
                print('Importing to database.')
                try:
                    # Convert dataframe to sql
                    event.listens_for(engine, 'before_cursor_execute')
                    sql = df.to_sql(name=table_name, con=engine, schema=schema, dtype=dtyp, index=False, chunksize=10000)
                    end_time = time.time()
                    print('Execution time: {:2f}'.format(end_time - start_time),'s')
                    # Close
                    engine.dispose()
                except ValueError:
                    engine.dispose()
                    display(dtyp)
                    time_elapsed = get_time_elapsed(time_elapsed)
                    raise
                except Exception as err:
                    print('Error: ' + str(err))
                    engine.dispose()
                    time_elapsed = get_time_elapsed(time_elapsed)
    else:
        print('Reading file as dataframe.')
        df = pd.read_excel(fname, sheet_name=sheet_name, index_col=False, engine='calamine')
        if sheet_name == None:
            for sheetnames in df.keys():
                try:
                    df[sheetnames].columns = df[sheetnames].columns.str.upper()
                    if len(df.keys()) > 1:
                        table_name = table_name+'_'+sheetnames
                    print('Pre-processing data types.')
                    if dtyp == None:
                        dtyp = get_sqlalchemy_dtyp_dict(df[sheetnames],engine)
                    else:
                        dtyp = dtyp
                    df[sheetnames] = df_fix_dtype(df[sheetnames],schema,dtyp=dtyp)
                    time_elapsed = get_time_elapsed(time_elapsed)
                    print('Fixing string datetime.')
                    df[sheetnames] = excel_df_fix_strftime(df[sheetnames])
                    time_elapsed = get_time_elapsed(time_elapsed)
                except Exception as err:
                    print('Error: ' + str(err))
                    engine.dispose()
                    time_elapsed = get_time_elapsed(time_elapsed)
                else:
                    print('Importing to database.')
                    try:
                        # Convert dataframe to sql
                        event.listens_for(engine, 'before_cursor_execute')
                        sql = df[sheetnames].to_sql(name=table_name, con=engine, schema=schema, dtype=dtyp, index=False, chunksize=10000)
                        end_time = time.time()
                        print('Execution time: {:2f}'.format(end_time - start_time),'s')
                        # Close
                        engine.dispose()
                    except ValueError:
                        engine.dispose()
                        display(dtyp)
                        time_elapsed = get_time_elapsed(time_elapsed)
                        raise
                    except Exception as err:
                        print('Error: ' + str(err))
                        engine.dispose()
                        time_elapsed = get_time_elapsed(time_elapsed)
        else:
            try:
                df.columns = df.columns.str.upper()
                print('Pre-processing data types.')
                if dtyp == None:
                    dtyp = get_sqlalchemy_dtyp_dict(df,engine)
                else:
                    dtyp = dtyp
                df = df_fix_dtype(df,schema,dtyp=dtyp)
                time_elapsed = get_time_elapsed(time_elapsed)
                print('Fixing string datetime.')
                df = excel_df_fix_strftime(df)
                time_elapsed = get_time_elapsed(time_elapsed)
            except Exception as err:
                print('Error: ' + str(err))
                engine.dispose()
                time_elapsed = get_time_elapsed(time_elapsed)
            else:
                print('Importing to database.')
                try:
                    # Convert dataframe to sql
                    event.listens_for(engine, 'before_cursor_execute')
                    sql = df.to_sql(name=table_name, con=engine, schema=schema, dtype=dtyp, index=False, chunksize=10000)
                    end_time = time.time()
                    print('Execution time: {:2f}'.format(end_time - start_time),'s')
                    # Close
                    engine.dispose()
                except ValueError:
                    engine.dispose()
                    display(dtyp)
                    time_elapsed = get_time_elapsed(time_elapsed)
                    raise
                except Exception as err:
                    print('Error: ' + str(err))
                    engine.dispose()
                    time_elapsed = get_time_elapsed(time_elapsed)

def df_to_sql(df,schema,table_name,dtyp=None,is_dtyp_fixed='NO'):
    import time
    start_time = time.time()
    time_elapsed = start_time
    
    # Set credential
    my_credentials = get_credentials('my_credentials.txt',schema)
    
    # Open
    engine = connect_sqlalchemy(my_credentials)
    
    try:
        df.columns = df.columns.str.upper()
        print('Pre-processing data types.')
        try:
            if dtyp == None:
                dtyp = get_sqlalchemy_dtyp_dict(df,engine)
                time_elapsed = get_time_elapsed(time_elapsed)
            else:
                dtyp = dtyp
            if is_dtyp_fixed.upper() == 'NO' or is_dtyp_fixed.upper() == 'N':
                df = df_fix_dtype(df,schema,dtyp=dtyp)
                time_elapsed = get_time_elapsed(time_elapsed)
                print('Fixing string datetime.')
                df = excel_df_fix_strftime(df)
                time_elapsed = get_time_elapsed(time_elapsed)
        except:
            engine.dispose()
    except Exception as err:
        print('Error: ' + str(err))
        engine.dispose()
        time_elapsed = get_time_elapsed(time_elapsed)
    else:
        print('Importing to database.')
        try:
            # Convert dataframe to sql
            for idx, (key,value) in enumerate(dtyp.items()):
                if 'VARCHAR' in str(value) and key == df.columns[idx]:
#                     df[key] = df[key].apply(lambda x: '' if is_nan(x) == True else str(x))
                    df[key] = df[key].replace('nan','')
            event.listens_for(engine, 'before_cursor_execute')
            '''
            col_sample is for testing Database Error DPI 1001: out of memory. When using pd.read_sql_query()
            and importing the generated dataframe back into database, there are no issues. However, when the
            generated dataframe is exported into a spreadsheet, re-opened via pandas, and imported back to
            database, there will be memory error due to large number of columns. Workaround will be to limit 
            chunksize parameter as small as possible or upgrade to higher RAM.
            fix it.
            '''
#             df_copy = df.copy()
#                         col_sample = []
#                         for idx,col in enumerate(df_copy.columns):
#                             if idx < 10:
#                                 col_sample.append(col)
#                         df_copy[col_sample]
#                         df = df_copy[col_sample]
            sql = df.to_sql(name=table_name, con=engine, schema=schema, dtype=dtyp, index=False, chunksize=10000)
            end_time = time.time()
            print('Execution time: {:2f}'.format(end_time - start_time),'s')
            # Close
            engine.dispose()
        except ValueError as err:
            engine.dispose()
            display(dtyp)
            time_elapsed = get_time_elapsed(time_elapsed)
        except Exception as err:
            print('Error: ' + str(err))
            engine.dispose()
            time_elapsed = get_time_elapsed(time_elapsed)
                    
def query_check(schema,table_name):
    query = f'''
                select count(*) num_cases
                from all_tab_columns
                where 1=1
                      and upper(owner) = '{str(schema.upper())}'
                      and upper(table_name) = '{str(table_name.upper())}'
              '''
    return query

def create_table_sql(query,fname,schema,table_name,sheet_name):
    start_time = time.time()
    # Set credential
    my_credentials = get_credentials('my_credentials.txt',schema)
    
    special_characters = "!@#$%^&*()[]{};:,./<>?\|`~-=+"
    
    
    fname_check_result = fname_check(fname,schema,table_name)
    is_file = fname_check_result['is_file']
    folder_path = fname_check_result['folder_path']
    fullpath = fname_check_result['fullpath']
    fname = fname_check_result['fname']
    
    if is_nan(table_name) == True:
        table_name = fname_check_result['table_name']
        table_name = clean_char(table_name,"_") # replace special characters with "_" and remove ending "_" character.
    else:
        table_name = table_name.upper()
        table_name = clean_char(table_name,"_") # replace special characters with "_" and remove ending "_" character.
                    
        
    print(f'''Query used: \n
-----------------------------------------------------------------
{query}
-----------------------------------------------------------------
''')
    if is_file == True:
        try:
            print('File already exists.')
            print('Fullpath:',fullpath)
        except Exception as err:
            print('Error: ' + str(err))
        else:
            # Open
            engine = connect_sqlalchemy(my_credentials)
            conn = engine.raw_connection()
            curr = conn.cursor()
            Session = sessionmaker(bind=engine)
            session = Session()
            
            try:
                df = pd.read_sql_query(query,engine)
            except Exception as err:
                print('Error: ' + str(err))
                curr.close
                engine.dispose()
            else:
                df.columns = df.columns.str.upper()
            
                dtyp = get_sqlalchemy_dtyp_dict(df,engine)

                # Check if table exists in database
                print('Checking in the database.')

                # Open file
                df_open = pd.read_excel(fname, sheet_name=None, engine='calamine')
                sheet_names = df_open.keys()
                for sheetnames in sheet_names:
                    table_name_base = table_name
                    if len(sheet_names) == 1:
                        curr.execute(query_check(schema,table_name_base))
                        if curr.fetchone()[0] >= 1:
                            is_new_table = input('Table already exists in the database. Do you want to create a new table? [YES/NO]  |  ')
                            if is_new_table.upper() == 'YES' or is_new_table.upper() == 'Y':
                                new_table_name = input('Insert new table name.  |  ')
                                new_table_name.upper()
                                new_table_name = clean_char(new_table_name,"_") # replace special characters with "_" and remove ending

                                new_fname = folder_path+'/'+new_table_name+''.join(pathlib.Path(fname).suffixes)
                                try:
                                    curr.execute(query_check(schema,new_table_name))
                                except Exception as err:
                                    print('Error: ' + str(err))
                                    curr.close()
                                    engine.dispose()
                                    break
                                else:
                                    if curr.fetchone()[0] >= 1:
                                        print('New table name already exists in the database.')
                                        if_exists_proceed = input('Do you want to replace the existing table with new values? [YES/NO]  |  ')
                                        if if_exists_proceed.upper() == 'YES' or if_exists_proceed.upper() == 'Y':
                                            # Reflect table and bind metadata to engine
                                            metadata = MetaData(conn)
                                            metadata.bind = engine
                                            reflected_table = Table(new_table_name, metadata, autoload_with=engine)
                                            # Drop the reflected table
                                            print('Dropping existing table.')
                                            reflected_table.drop(checkfirst=True)
                                            session.commit()
                                            # Convert dataframe to sql
                                            print('Importing to database.')
                                            event.listens_for(engine, 'before_cursor_execute')
                                            try:
                                                sql = df.to_sql(name=new_table_name, con=engine, schema=schema, dtype=dtyp, index=False)
                                            except Exception as err:
                                                print('Error: ' + str(err))
                                                curr.close()
                                                engine.dispose()
                                            else:
                                                print('Table replaced.')
                                        else:
                                            print('No action done.')
                                    else:
                                        print('New fullpath:',new_fname.replace('/','\\'))
                                        # Convert dataframe to excel and save
                                        if fname.endswith('.xlsx') or fname.endswith('.xlsm') or fname.endswith('.xls') or fname.endswith('.csv'):
                                            df_copy = df.copy()
                                            df_to_excel(df_copy,schema,new_fname,sheetnames,dtyp=dtyp)
                                            del df_copy
                                        curr.close()
                                        # Convert dataframe to sql
                                        print('Importing to database.')
                                        event.listens_for(engine, 'before_cursor_execute')
                                        try:
                                            sql = df.to_sql(name=new_table_name, con=engine, schema=schema, dtype=dtyp, index=False)
                                        except Exception as err:
                                            print('Error: ' + str(err))
                                            curr.close()
                                            engine.dispose()
                                        else:
                                            print(f"Table [{new_table_name}] created.")
                            elif is_new_table.upper() == 'NO' or is_new_table.upper() == 'N':
                                if_exists_proceed = input('Do you want to replace the existing table? [YES/NO]')
                                if if_exists_proceed.upper() == 'YES' or if_exists_proceed.upper() == 'Y':
                                    # Reflect table and bind metadata to engine
                                    metadata = MetaData(conn)
                                    metadata.bind = engine
                                    reflected_table = Table(table_name, metadata, autoload_with=engine)
                                    # Drop the reflected table
                                    print('Dropping existing table.')
                                    reflected_table.drop(checkfirst=True)
                                    session.commit()
                                    # Convert dataframe to sql
                                    print('Importing to database.')
                                    event.listens_for(engine, 'before_cursor_execute')
                                    try:
                                        sql = df.to_sql(name=table_name, con=engine, schema=schema, dtype=dtyp, index=False)
                                    except Exception as err:
                                        print('Error: ' + str(err))
                                        curr.close()
                                        engine.dispose()
                                    else:
                                        print('Table replaced.')
                                else:
                                    print('No action done.')
                            else:
                                print('Invalid input. Ending prompt.')
                        else:
                            print('No existing table. Importing to database.')
                            table_name.upper()
                            table_name = clean_char(table_name,"_") # replace special characters with "_" and remove ending
                            # Convert dataframe to sql
                            event.listens_for(engine, 'before_cursor_execute')
                            try:
                                sql = df.to_sql(name=table_name, con=engine, schema=schema, dtype=dtyp, index=False)
                            except Exception as err:
                                print('Error: ' + str(err))
                                curr.close()
                                engine.dispose()
                            else:
                                print(f"Table [{table_name}] created.")
                    else:
                        table_name_sheet = table_name_base + '_' +  sheetnames
                        table_name_sheet.upper()
                        table_name_sheet = clean_char(table_name_sheet,"_") # replace special characters with "_" and remove ending
                        curr.execute(query_check(schema,table_name_sheet))
                        if curr.fetchone()[0] >= 1:
                            query_column = f'''
                                                select
                                                       column_name
                                                      ,count(*) num_cases
                                                from all_tab_columns
                                                where 1=1
                                                      and upper(owner) = '{str(schema.upper())}'
                                                      and upper(table_name) = '{str(table_name.upper())}'
                                                group by
                                                       column_name
                                            '''
                            df_query = pd.read_sql_query(query_column,engine)
                            df_query = df_query.drop('num_cases', axis=1)

                            df_col = df.columns.str.upper()
                            df_exist = pd.DataFrame({'column_name': df_col})

                            # Check if any value in df1['col1'] does not exist in df2['col1']
                            missing_values = df_exist[~df_exist['column_name'].isin(df_query['column_name'])]

                            if not missing_values.empty:
                                is_new_table = input('Table already exists in the database. Do you want to create a new table? [YES/NO]  |  ')
                                if is_new_table.upper() == 'YES' or is_new_table.upper() == 'Y':
                                    new_table_name = input('Insert new table name.  |  ')
                                    new_table_name.upper()
                                    new_table_name = clean_char(new_table_name,"_") # replace special characters with "_" and remove ending

                                    new_fname = folder_path+'/'+new_table_name+''.join(pathlib.Path(fname).suffixes)
                                    try:
                                        curr.execute(query_check(schema,new_table_name))
                                    except Exception as err:
                                        print('Error: ' + str(err))
                                        curr.close()
                                        engine.dispose()
                                    else:
                                        if curr.fetchone()[0] >= 1:
                                            print('New table name already exists in the database.')
                                            if_exists_proceed = input('Do you want to replace the existing table with new values? [YES/NO]  |  ')
                                            if if_exists_proceed.upper() == 'YES' or if_exists_proceed.upper() == 'Y':
                                                # Reflect table and bind metadata to engine
                                                metadata = MetaData(conn)
                                                metadata.bind = engine
                                                reflected_table = Table(new_table_name, metadata, autoload_with=engine)
                                                # Drop the reflected table
                                                print('Dropping existing table.')
                                                reflected_table.drop(checkfirst=True)
                                                session.commit()
                                                # Convert dataframe to sql
                                                print('Importing to database.')
                                                event.listens_for(engine, 'before_cursor_execute')
                                                try:
                                                    sql = df.to_sql(name=new_table_name, con=engine, schema=schema, dtype=dtyp, index=False)
                                                except Exception as err:
                                                    print('Error: ' + str(err))
                                                    curr.close()
                                                    engine.dispose()
                                                else:
                                                    print('Table replaced.')
                                            else:
                                                print('No action done.')
                                        else:
                                            print('New fullpath:',new_fname.replace('/','\\'))
                                            # Convert dataframe to excel and save
                                            if fname.endswith('.xlsx') or fname.endswith('.xlsm') or fname.endswith('.xls') or fname.endswith('.csv'):
                                                df_copy = df.copy()
                                                df_to_excel(df_copy,schema,new_fname,table_name_sheet,dtyp=dtyp)
                                                del df_copy
                                            curr.close()
                                            # Convert dataframe to sql
                                            print('Importing to database.')
                                            event.listens_for(engine, 'before_cursor_execute')
                                            try:
                                                sql = df.to_sql(name=new_table_name, con=engine, schema=schema, dtype=dtyp, index=False)
                                            except Exception as err:
                                                print('Error: ' + str(err))
                                                curr.close()
                                                engine.dispose()
                                            else:
                                                print(f"Table [{new_table_name}] created.")
                                elif is_new_table.upper() == 'NO' or is_new_table.upper() == 'Y':
                                    if_exists_proceed = input('Do you want to replace the existing table? [YES/NO]  |  ')
                                    if if_exists_proceed.upper() == 'YES' or if_exists_proceed.upper() == 'Y':
                                        # Reflect table and bind metadata to engine
                                        metadata = MetaData(conn)
                                        metadata.bind = engine
                                        reflected_table = Table(table_name_sheet, metadata, autoload_with=engine)
                                        # Drop the reflected table
                                        print('Dropping existing table.')
                                        reflected_table.drop(checkfirst=True)
                                        session.commit()
                                        # Convert dataframe to sql
                                        print('Importing to database.')
                                        event.listens_for(engine, 'before_cursor_execute')
                                        try:
                                            sql = df.to_sql(name=table_name_sheet, con=engine, schema=schema, dtype=dtyp, index=False)
                                        except Exception as err:
                                            print('Error: ' + str(err))
                                            curr.close()
                                            engine.dispose()
                                        else:
                                            print('Table replaced.')
                                    else:
                                        print('No action done.')
                                else:
                                    print('Invalid input. Ending prompt.')

                            else:
                                is_new_table = input('Identical table with same table_name and column_names is found in the database. Do you want to create a new table? [YES/NO]  |  ')
                                if is_new_table.upper() == 'YES' or is_new_table.upper() == 'Y':
                                    new_table_name = input('Insert new table name.  |  ')
                                    new_table_name.upper()
                                    new_table_name = clean_char(new_table_name,"_") # replace special characters with "_" and remove ending "_" character.

                                    new_fname = folder_path+'/'+new_table_name+''.join(pathlib.Path(fname).suffixes)
                                    try:
                                        curr.execute(query_check(schema,new_table_name))
                                    except Exception as err:
                                        print('Error: ' + str(err))
                                        curr.close()
                                        engine.dispose()
                                    else:
                                        if curr.fetchone()[0] >= 1:
                                            print('New table name already exists in the database.')
                                            if_exists_proceed = input('Do you want to replace the existing table with new values? [YES/NO]  |  ')
                                            if if_exists_proceed.upper() == 'YES' or if_exists_proceed.upper() == 'Y':
                                                # Reflect table and bind metadata to engine
                                                metadata = MetaData(conn)
                                                metadata.bind = engine
                                                reflected_table = Table(new_table_name, metadata, autoload_with=engine)
                                                # Drop the reflected table
                                                print('Dropping existing table.')
                                                reflected_table.drop(checkfirst=True)
                                                session.commit()
                                                # Convert dataframe to sql
                                                print('Importing to database.')
                                                event.listens_for(engine, 'before_cursor_execute')
                                                try:
                                                    sql = df.to_sql(name=new_table_name, con=engine, schema=schema, dtype=dtyp, index=False)
                                                except Exception as err:
                                                    print('Error: ' + str(err))
                                                    curr.close()
                                                    engine.dispose()
                                                else:
                                                    print('Table replaced.')
                                            else:
                                                print('No action done.')
                                        else:
                                            print('New fullpath:',new_fname.replace('/','\\'))
                                            # Convert dataframe to excel and save
                                            if fname.endswith('.xlsx') or fname.endswith('.xlsm') or fname.endswith('.xls') or fname.endswith('.csv'):
                                                df_copy = df.copy()
                                                df_to_excel(df_copy,schema,new_fname,table_name_sheet,dtyp=dtyp)
                                                del df_copy
                                            curr.close()                                            
                                            # Convert dataframe to sql
                                            print('Importing to database.')
                                            event.listens_for(engine, 'before_cursor_execute')
                                            try:
                                                sql = df.to_sql(name=new_table_name, con=engine, schema=schema, dtype=dtyp, index=False)
                                            except Exception as err:
                                                print('Error: ' + str(err))
                                                curr.close()
                                                engine.dispose()
                                            else:
                                                print(f"Table [{new_table_name}] created.")
                                elif is_new_table.upper() == 'NO' or is_new_table.upper() == 'N':
                                    if_exists_proceed = input('Do you want to replace the existing table? [YES/NO]  |  ')
                                    if if_exists_proceed.upper() == 'YES' or if_exists_proceed.upper() == 'Y':
                                        # Reflect table and bind metadata to engine
                                        metadata = MetaData(conn)
                                        metadata.bind = engine
                                        reflected_table = Table(table_name, metadata, autoload_with=engine)
                                        # Drop the reflected table
                                        print('Dropping existing table.')
                                        reflected_table.drop(checkfirst=True)
                                        session.commit()
                                        # Convert dataframe to sql
                                        print('Importing to database.')
                                        event.listens_for(engine, 'before_cursor_execute')
                                        try:
                                            sql = df.to_sql(name=table_name, con=engine, schema=schema, dtype=dtyp, index=False)
                                        except Exception as err:
                                            print('Error: ' + str(err))
                                            curr.close()
                                            engine.dispose()
                                        else:
                                            print('Table replaced.')
                                    else:
                                        print('No action done.')
                                else:
                                    print('Invalid input. Ending prompt.')
                        else:
                            print('No existing table. Importing to database.')
                            try:
                                sql = df[sheetnames].to_sql(name=table_name_sheet, con=engine, schema=schema, dtype=dtyp, index=False)
                            except Exception as err:
                                print('Error: ' + str(err))
                                curr.close()
                                engine.dispose()
                            else:
                                print(f"Table [{table_name}] created.")

                # Close
                engine.dispose()
                print('Finished.')
    else:
        print('No matching file found.')
        # Open
        engine = connect_sqlalchemy(my_credentials)
        conn = engine.raw_connection()
        curr = conn.cursor()
        Session = sessionmaker(bind=engine)
        session = Session()
        if sheet_name == None:
            sheet_name = 'Sheet1'
        try:
            df = pd.read_sql_query(query,engine)
        except Exception as err:
            print('Error: ' + str(err))
            engine.dispose()
        else:
            df.columns = df.columns.str.upper()

            dtyp = get_sqlalchemy_dtyp_dict(df,engine)

            # Check if table exists in database
            print('Checking in the database.')

            try:
                curr.execute(query_check(schema,table_name))
            except Exception as err:
                print('Error: ' + str(err))
                curr.close()
                engine.dispose()
            else:
                if curr.fetchone()[0] >= 1:
                    query_column = f'''
                                        select
                                               column_name
                                              ,count(*) num_cases
                                        from all_tab_columns
                                        where 1=1
                                              and upper(owner) = '{str(schema.upper())}'
                                              and upper(table_name) = '{str(table_name.upper())}'
                                        group by
                                               column_name
                                    '''
                    df_query = pd.read_sql_query(query_column,engine)
                    df_query = df_query.drop('num_cases', axis=1)

                    df_col = df.columns.str.upper()
                    df_exist = pd.DataFrame({'column_name': df_col})

                    # Check if any value in df1['col1'] does not exist in df2['col1']
                    missing_values = df_exist[~df_exist['column_name'].isin(df_query['column_name'])]

                    if not missing_values.empty:
                        is_new_table = input('Table already exists in the database. Do you want to create a new table? [YES/NO]  |  ')
                        if is_new_table.upper() == 'YES' or is_new_table.upper() == 'Y':
                            new_table_name = input('Insert new table name.  |  ')
                            new_table_name.upper()
                            new_table_name = clean_char(new_table_name,"_") # replace special characters with "_" and remove ending "_" character.

                            new_fname = folder_path+'/'+new_table_name+''.join(pathlib.Path(fname).suffixes)
                            try:
                                curr.execute(query_check(schema,new_table_name))
                            except Exception as err:
                                print('Error: ' + str(err))
                                curr.close()
                                engine.dispose()
                            else:
                                if curr.fetchone()[0] >= 1:
                                    print('New table name already exists in the database.')
                                    is_replace = input('Do you want to replace the existing table with new values? [YES/NO]  |  ')
                                    if is_replace.upper() == 'YES' or is_replace.upper() == 'Y':
                                        # Reflect table and bind metadata to engine
                                        metadata = MetaData(conn)
                                        metadata.bind = engine
                                        reflected_table = Table(new_table_name, metadata, autoload_with=engine)
                                        # Drop the reflected table
                                        print('Dropping existing table.')
                                        reflected_table.drop(checkfirst=True)
                                        session.commit()
                                        print('Importing to database.')
                                        event.listens_for(engine, 'before_cursor_execute')
                                        try:
                                            sql = df.to_sql(name=new_table_name, con=engine, schema=schema, dtype=dtyp, index=False)
                                        except Exception as err:
                                            print('Error: ' + str(err))
                                            curr.close()
                                            engine.dispose()
                                        else:
                                            print('Table replaced.')
                                    else:
                                        print('No action done.')
                                else:
                                    print('New fullpath:',new_fname)
                                    # Convert dataframe to excel and save
                                    if fname.endswith('.xlsx') or fname.endswith('.xlsm') or fname.endswith('.xls') or fname.endswith('.csv'):
                                        df_copy = df.copy()
                                        df_to_excel(df_copy,schema,new_fname,sheet_name,dtyp=dtyp)
                                        del df_copy
                                    curr.close()
                                    # Convert dataframe to sql
                                    print('Importing to database.')
                                    event.listens_for(engine, 'before_cursor_execute')
                                    try:
                                        sql = df.to_sql(name=new_table_name, con=engine, schema=schema, dtype=dtyp, index=False)
                                    except Exception as err:
                                        print('Error: ' + str(err))
                                        curr.close()
                                        engine.dispose()
                                    else:
                                        print(f"Table [{new_table_name}] created.")
                        elif is_new_table.upper() == 'NO' or is_new_table.upper() == 'N':
                            if_exists_proceed = input('Do you want to replace the existing table? [YES/NO]')
                            if if_exists_proceed.upper() == 'YES' or if_exists_proceed.upper() == 'Y':
                                # Reflect table and bind metadata to engine
                                metadata = MetaData(conn)
                                metadata.bind = engine
                                reflected_table = Table(table_name, metadata, autoload_with=engine)
                                # Drop the reflected table
                                print('Dropping existing table.')
                                reflected_table.drop(checkfirst=True)
                                session.commit()
                                # Convert dataframe to sql
                                print('Importing to database.')
                                event.listens_for(engine, 'before_cursor_execute')
                                try:
                                    sql = df.to_sql(name=table_name, con=engine, schema=schema, dtype=dtyp, index=False)
                                except Exception as err:
                                    print('Error: ' + str(err))
                                    curr.close()
                                    engine.dispose()
                                else:
                                    print('Table replaced.')
                        else:
                            print('Invalid input. Ending prompt.')

                    else:
                        is_new_table = input('Identical table with same table_name and column_names is found in the database. Do you want to create a new table? [YES/NO]  |  ')
                        if is_new_table.upper() == 'YES' or is_new_table.upper() == 'Y':
                            new_table_name = input('Insert new table name.  |  ')
                            new_table_name.upper()
                            new_table_name = clean_char(new_table_name,"_") # replace special characters with "_" and remove ending "_" character.

                            new_fname = folder_path+'/'+new_table_name+''.join(pathlib.Path(fname).suffixes)
                            try:
                                curr.execute(query_check(schema,new_table_name))
                            except Exception as err:
                                print('Error: ' + str(err))
                                curr.close()
                                engine.dispose()
                            else:
                                if curr.fetchone()[0] >= 1:
                                    print('New table name already exists in the database.')
                                    if_exists_proceed = input('Do you want to replace the existing table with new values? [YES/NO]  |  ')
                                    if if_exists_proceed.upper() == 'YES' or if_exists_proceed.upper() == 'Y':
                                        # Reflect table and bind metadata to engine
                                        metadata = MetaData(conn)
                                        metadata.bind = engine
                                        reflected_table = Table(new_table_name, metadata, autoload_with=engine)
                                        # Drop the reflected table
                                        print('Dropping existing table.')
                                        reflected_table.drop(checkfirst=True)
                                        session.commit()
                                        # Convert dataframe to sql
                                        print('Importing to database.')
                                        event.listens_for(engine, 'before_cursor_execute')
                                        try:
                                            sql = df.to_sql(name=new_table_name, con=engine, schema=schema, dtype=dtyp, index=False)
                                        except Exception as err:
                                            print('Error: ' + str(err))
                                            curr.close()
                                            engine.dispose()
                                        else:
                                            print('Table replaced.')
                                    else:
                                        print('No action done.')
                                else:
                                    print('New fullpath:',new_fname.replace('/','\\'))
                                    # Convert dataframe to excel and save
                                    if fname.endswith('.xlsx') or fname.endswith('.xlsm') or fname.endswith('.xls') or fname.endswith('.csv'):
                                        df_copy = df.copy()
                                        df_to_excel(df_copy,schema,new_fname,sheet_name,dtyp=dtyp)
                                        del df_copy
                                    curr.close()
                                    # Convert dataframe to sql
                                    print('Importing to database.')
                                    event.listens_for(engine, 'before_cursor_execute')
                                    try:
                                        sql = df.to_sql(name=new_table_name, con=engine, schema=schema, dtype=dtyp, index=False)
                                    except Exception as err:
                                        print('Error: ' + str(err))
                                        curr.close()
                                        engine.dispose()
                                    else:
                                        print(f"Table [{new_table_name}] created.")
                        elif is_new_table.upper() == 'NO' or is_new_table.upper() == 'N':
                            if_exists_proceed = input('Do you want to replace the existing table? [YES/NO]  |  ')
                            if if_exists_proceed.upper() == 'YES' or if_exists_proceed.upper() == 'Y':
                                # Reflect table and bind metadata to engine
                                metadata = MetaData(conn)
                                metadata.bind = engine
                                reflected_table = Table(table_name, metadata, autoload_with=engine)
                                # Drop the reflected table
                                print('Dropping existing table.')
                                reflected_table.drop(checkfirst=True)
                                session.commit()
                                # Convert dataframe to sql
                                print('Importing to database.')
                                event.listens_for(engine, 'before_cursor_execute')
                                try:
                                    sql = df.to_sql(name=table_name, con=engine, schema=schema, dtype=dtyp, index=False)
                                except Exception as err:
                                    print('Error: ' + str(err))
                                    curr.close()
                                    engine.dispose()
                                else:
                                    print('Table replaced.')
                            else:
                                print('No action done.')
                        else:
                            print('Invalid input. Ending prompt.')
                else:
                    # Convert dataframe to excel and save
                    if fname.endswith('.xlsx') or fname.endswith('.xlsm') or fname.endswith('.xls') or fname.endswith('.csv'):
                        df_copy = df.copy()
                        df_to_excel(df_copy,schema,fname,sheet_name,dtyp=dtyp)
                        del df_copy
                    # Convert dataframe to sql
                    print('No existing table. Importing to database.')
                    print(fname)
                    event.listens_for(engine, 'before_cursor_execute')
                    try:
                        sql = df.to_sql(name=table_name, con=engine, schema=schema, dtype=dtyp, index=False)
                    except Exception as err:
                        print('Error: ' + str(err))
                        curr.close()
                        engine.dispose()
                    else:
                        print(f"Table [{table_name}] created.")
        
            # Close
            engine.dispose()
            print('Finished.')
        
    end_time = time.time()
    print('Execution time: {:2f}'.format(end_time - start_time),'s')



#--------------------------------- Data Visualization


def appshare_table_and_plot(data, gb_key, col_time, list_apps, display_table = True, display_plot = True):
    
    share_df = data.copy()
    
    aggs = {gb_key : ['count']}
    for app in list_apps:
        aggs[app] = ['sum', 'mean']
    
    share_df = share_df.groupby(col_time).agg(aggs)
    agg_col_names = ['{}_{}'.format(x[1], x[0]) for x in share_df.columns.tolist()]
    agg_col_names[0] = 'count samples'
    
    list_new_columns = []
    for col in agg_col_names:
        if col.startswith('sum'):
            new_col = 'count has ' + col.split('sum_')[1]
        elif col.startswith('mean'):
            new_col = 'percentage has ' + col.split('mean_')[1]
        else:
            new_col = col
        
        list_new_columns.append(new_col)
    
    share_df.columns = pd.Index(list_new_columns)
    
    for col in list(share_df.columns):
        if col.startswith('percentage has '):
            share_df[col] = round(share_df[col] * 100, 2)
    
    if display_table:
        display(share_df)
        print('')
    
    if display_plot:
        
        list_col_to_plot = ['percentage has ' + col for col in list_apps]
        plot_df = share_df[list_col_to_plot].copy()
        
        plot_df.plot(figsize = (15, 8), linewidth = 2, marker = 'o', markersize = 8, cmap = 'Paired')
        
        plt.title('Percentage of certain apps per ' + col_time, fontsize = 16)
        plt.xlabel(col_time, fontsize = 14)
        plt.ylabel('Percentage', fontsize = 14)
        plt.legend(loc = 'best')
        plt.show()
        print('')
        
        
def multiple_barplot(list_df, list_title, list_x_axis, list_y_axis, num_cols, diff_colors = False, c_palette = None):
    
    num_df = len(list_df)
    num_rows = int(np.ceil(num_df / num_cols))
    
    horizontal = 14 * num_cols
    vertical = 8 * num_rows
    
    plt.figure(figsize = (horizontal, vertical))
    
    gs = gridspec.GridSpec(num_rows, num_cols)
    gs.update(wspace = 0.2, hspace = 0.5)
    
    list_ax = []
    row = 0
    col = 0
    
    for i in range(0, num_df):
        if col < num_cols - 1:
            ax = plt.subplot(gs[row, col])
            list_ax.append(ax)
            col = col + 1
        elif col == num_cols -1:
            ax = plt.subplot(gs[row, col])
            list_ax.append(ax)
            col = 0
            row = row + 1
            
    list_color = ['#7e1e9c', '#15b01a', '#0343df', '#ff81c0', '#653700', '#e50000', '#029386'
             , '#f97306', '#929591', '#033500' , '#06c2ac' , '#dbb40c', '#c04e01', '#000000'
             , '#0652ff', '#1fa774', '#7FFF00', '#292421']

    c_mult = int(np.ceil(num_df / len(list_color)))
    final_list_color = list_color * c_mult
    final_list_color = final_list_color[:num_df]
    
    for df, ax, title, x_axis, y_axis, color in zip(list_df, list_ax, list_title, list_x_axis, list_y_axis, final_list_color):
        if diff_colors:
            sns.barplot(df[x_axis], df[y_axis], color = color, ax = ax)
        elif c_palette is not None:
            sns.barplot(df[x_axis], df[y_axis], palette = c_palette, ax = ax)
        else:
            sns.barplot(df[x_axis], df[y_axis], color = 'darkblue', ax = ax)
        
        ax.set_title(title, fontsize = 20)
        ax.set_xlabel(x_axis, fontsize = 15)
        ax.set_ylabel(y_axis, fontsize = 15)
        ax.tick_params(axis = 'both', labelsize = 13)
        ax.tick_params(axis = 'x', rotation = 90)



#--------------------------------- Hadoop/PySpark scripts below

#-------- spark = SparkSession.builder.config('spark.sql.session.timeZone', 'Asia/Manila').getOrCreate()


def get_engine(filename=None):
    """
    Creates SQLAlchemy engine (as singleton) based on connection string from ~/.ora_cnn or filename if specified.

    :param filename: name of file with connection string
    :return: SQLAlchemy engine
    """
    global _engine
    if filename is None:
        filename=os.path.join(os.path.expanduser('~'), '.ora_cnn')
    need_reconnect = False
    if '_engine' in globals() and _engine is not None:
        try: 
            pd.read_sql_query('select * from dual', con=_engine)
        except sqlalchemy.exc.DatabaseError as e:           
            warnings.warn(traceback.format_exc())
            need_reconnect = True
    if need_reconnect or '_engine' not in globals() or _engine is None:  
        with open(filename) as f:
            ora_cnn = f.read()
        _engine = create_engine(ora_cnn, echo=False, encoding='utf-8')
    return _engine


def to_sql(df, name, con = None, chunksize=10000, if_exists='replace', **kwargs):
    """
    Writes pandas.DataFrame to database.

    See params of pandas.DataFrame.to_sql. This function generate proper column names (they should less than
    30 characters). And fixes performance issues by using varchar instead of BLOB used by default.
    """
    if con is None:
        con = get_engine()
    index_cols = [c[:30].lower() for c in (df.index.names if type(df.index) == pd.MultiIndex else [df.index.name]
                  if df.index.name is not None else [])]
    
    if sys.version_info[0] < 3:
        columns=[unicode(c[:30].lower()) for c in df.columns]
    else:
        columns=[c[:30].lower() for c in df.columns]

    if len(set(index_cols+columns)) != len(index_cols+columns):
        raise ValueError('Index/column names are not unique after truncation to 30 characters and converting to '
                         'lowercase')

    df_copy = df.rename(columns=dict(zip(df.columns, columns)))
    if len(index_cols) == 1:
        df_copy.index.rename(index_cols[0], inplace=True)
    elif len(index_cols) > 1:
        df_copy.index.rename(index_cols, inplace=True)

    dtyp = dict([(i,types.VARCHAR(i.str.len().max())) for i in
                 ([ii.name for ii in df_copy.index.levels] if type(df_copy.index) == pd.MultiIndex
                  else [df_copy.index] if df_copy.index.name is not None else []) if i.dtype == 'object' or
                  i.dtype.name=='category'] +
                  [(c,types.VARCHAR(df_copy[c].str.len().max())) for c in df_copy.columns
                   if df_copy[c].dtype == 'object' or df_copy[c].dtype.name == 'category'])
    df_copy[columns].to_sql(name, con=con, chunksize=chunksize, if_exists=if_exists, dtype=dtyp, **kwargs)


def read_sql(sql, con = None, refresh=True, optimize_types=False, str2cat=True, **kwargs):
    """
    Caching data read from database to csv (use refresh=True to refresh case)
    Optimizing data types (optimize_types=True)

    :param sql: SQL query
    :param con: existing SQLAlchemy connection, will be creaye via get_engine if None passed
    :param refresh: should cache refreshed from re-reading from database
    :param optimize_types: should type optimization applied
    :param str2cat: convert string columns to pandas.category

    :return: pandas.DataFrame

    """
    if con is None:
        con = get_engine()
    dir = 'db_cache'
    if not os.path.exists(dir):
        os.makedirs(dir)
    if sys.version_info[0] < 3:
        hash_filename= sha512(sql).hexdigest()
    else:
        hash_filename= sha512(sql.encode('utf8')).hexdigest()    

    path=os.path.join(dir, hash_filename)
    if os.path.exists(path) and not refresh:
        df = pd.read_pickle(path)
        return df
        
    df = pd.read_sql_query(sql, con=con, **kwargs)
    if optimize_types:
        # types optimization works only for first reading from database or for regresh=True !!!
        new_types = get_optimized_types(df, str2cat=str2cat)
        for x in new_types.items():
            df[x[0]] = df[x[0]].astype(x[1])

    df.to_pickle(path)
    return df


def read_csv(filepath_or_buffer, optimize_types=False, sample_nrows=None, str2cat=True, minimalfloat='float64', **kwargs):
    """
    Reading CSV with types optimization.

    :param filepath_or_buffer: file path ot file object
    :param optimize_types: do types optimization
    :param sample_nrows: number of rows to use for type optimization, all if None
    :param str2cat: do conversion of str to pandas.Category
    :param minimalfloat: minimal bit size of float 

    :return: pandas.DataFrame
    """
    if not optimize_types:
        return pd.read_csv(filepath_or_buffer, **kwargs)

    df = pd.read_csv(filepath_or_buffer, nrows=sample_nrows, **kwargs)
    new_types = get_optimized_types(df, str2cat=str2cat, minimalfloat=minimalfloat)

    if sample_nrows is not None: # subsample based types inference - takes less memory
        df = pd.read_csv(filepath_or_buffer, dtype=new_types, **kwargs)
    else: # types optimization AFTER reading entire DataFrame
        for x in new_types.items():
            df[x[0]] = df[x[0]].astype(x[1])
    return df


def get_optimized_types(df, str2cat=True, minimalfloat='float64'):
    """
    Detects "minimum" types of DataFrame columns.

    :param df: DataFrame
    :param str2cat: should str be converted to pandas.Category
    :param minimalfloat: minimal bit size of float 


    :return: dict of column name -> type
    """
    new_types = {}
    for x in df.dtypes.iteritems():
        if x[1].name.startswith('int') or x[1].name.startswith('uint'):
            flag = False
            for t in ['uint8', 'int8', 'uint16', 'int16', 'uint32', 'int32', 'uint64', 'int64']:
                if df[x[0]].min() >= np.iinfo(t).min and df[x[0]].max() <= np.iinfo(t).max:
                    new_types[x[0]] = t
                    flag = True
                    break
            if not flag:
                raise ValueError()
        elif x[1].name.startswith('float'):
            new_t = x[1]
            # restric possible float sizes, using float64 will result in empty list
            # and no attempts for smaller floats will occur
            possible_floats = [f for f in ['float32', 'float16'] if int(f[-2:]) >= int(minimalfloat[-2:])]
            for t in possible_floats:
                unique_num = df[x[0]].unique().shape[0]
                if df[x[0]].astype(t).unique().shape[0] < unique_num:
                    break
                new_t = t
            new_types[x[0]] = new_t
        elif str2cat and x[1] == 'O' and df[x[0]].apply(lambda x: type(x) == str or isinstance(x, numbers.Number) and
                                                        np.isnan(x) or x is None).all():
            new_types[x[0]] = 'category'
    return new_types



def get_optimal_numerical_type(series, float_type='float32'):
    """
    Returns optimal numerical type for pd.Series.

    Integer types will be chosen based on min/max values. Unsigned version if Series contains no negative numbers.
    Float32 is default unless converting to it would lower number of unique values in the Series. In that case float64 will be used.

    :param series: Pandas series
    :param float_type: can be used to force float64 

    :return: optimal dtype as string

    :raises: ValueError when series is no numerical
    """
    if pd.api.types.is_numeric_dtype(series.values.dtype):
        pd_type = pd.to_numeric(series).dtype.name
    else:
        raise ValueError('Series \'{0}\':[{1}] is not numerical.'.format(series.name, series.dtype))

    if 'int' in pd_type:
        if series.min() >= 0:
            for t in ['uint8', 'uint16', 'uint32', 'uint64']:
                if series.max() < np.iinfo(t).max:
                    break
        else:
            for t in ['int8', 'int16', 'int32', 'int64']:
                if series.max() < np.iinfo(t).max and series.min() > np.iinfo(t).min:
                    break
    else:
        if series.astype(np.float32).nunique() == series.nunique() and float_type=='float32':
            t = 'float32'
        else:
            t = 'float64'
    return t


# Function that will mainly do the actual cleansing parsing of the data
def cleanse_and_parse_func(data, list_toparse):
    
    for col in list_toparse:
        data = data.withColumn(col, regexp_replace(col,"\[",""))
        data = data.withColumn(col, regexp_replace(col,"\]",""))
        data = data.withColumn(col, regexp_replace(col,'"',"")) #Replace double quotes
        
        
    list_df = []
    
    for col in list_toparse:
        temp_df = data.withColumn(col, explode(split(data[col],",")))
        temp_df = temp_df.withColumn("id", monotonically_increasing_id())
        list_df.append(temp_df)
        
    
    schema = StructType([StructField('id', StringType(), True)\
                    , StructField('id_batch', IntegerType(), True)\
                    , StructField('dtime_datascore', TimestampType(), True)\
                    , StructField('part_date', IntegerType(), True)])

    df = spark.createDataFrame(sc.emptyRDD(), schema)
    
    
    for df_temp in list_df:
        df = df.union(df_temp.select('id', 'id_batch', 'dtime_datascore', 'part_date'))
        
    df = df.select('id', 'id_batch', 'dtime_datascore', 'part_date').distinct()
    
    
    list_df_new = []
    
    for df_temp, col in zip(list_df, list_toparse):
        df_temp = df_temp.select(col, 'id')
        list_df_new.append(df_temp)
    
        
    for df_temp in list_df_new:
        df = df.join(df_temp, on = 'id', how = 'left')
        
    
    df = df.select(list_toparse + ['id_batch', 'dtime_datascore', 'part_date'])
    
    return df



# Importing the functions that will be used in parsing the different categories of datascore

def parse_datascoreBatches(whole_data_hive):
    
    
    start_time = time.time()
    print('Currently parsing datascoreBatches...')
    
    whole_data_hive.createOrReplaceTempView('test_bigdata_data_hive_temp')
     
    datascoreBatches = spark.sql(
        '''
            select
                   id_batch
                  ,dtime_datascore
                  ,part_date
            from test_bigdata_data_hive_temp
        '''
        )
    
    datascoreBatches.repartition(200).write.format('parquet').partitionBy('part_date').mode('append').saveAsTable('TEST_BIGDATA_OWNER.DATASCORE_BATCHES')
    
    print('Execution time for datascoreBatches:', time.time() - start_time)
    print('')


def parse_applicationInfoList(whole_data_hive):
    
    start_time = time.time()
    print('Currently parsing applicationInfoList...')
    
    whole_data_hive.createOrReplaceTempView('test_bigdata_data_hive_temp')
    
    applicationInfoList_df = spark.sql(
    '''
        select
               id_batch
              ,dtime_datascore
              ,part_date
              ,get_json_object(content, "$.applicationInfoList[*].packageName") as packageName
              ,get_json_object(content, "$.applicationInfoList[*].firstInstallTime") as firstInstallTime
              ,get_json_object(content, "$.applicationInfoList[*].lastUpdateTime") as lastUpdateTime
              ,get_json_object(content, "$.applicationInfoList[*].versionName") as versionName
              ,get_json_object(content, "$.applicationInfoList[*].versionCode") as versionCode
              ,get_json_object(content, "$.applicationInfoList[*].flags") as flags
              ,get_json_object(content, "$.applicationInfoList[*].baseRevisionCode") as baseRevisionCode
              ,get_json_object(content, "$.applicationInfoList[*].installLocation") as installLocation
              ,get_json_object(content, "$.applicationInfoList[*].sharedUserId") as sharedUserId
              ,get_json_object(content, "$.applicationInfoList[*].splitNames") as splitNames
              ,get_json_object(content, "$.applicationInfoList[*].splitRevisionCodes") as splitRevisionCodes
        from test_bigdata_data_hive_temp
    '''
    )

    col_toparse = ['packageName', 'firstInstallTime', 'lastUpdateTime', 'versionName', 'versionCode', 'flags', 'baseRevisionCode', 'installLocation'
                    , 'sharedUserId', 'splitNames', 'splitRevisionCodes']
                    
    applicationInfoList_df = cleanse_and_parse_func(applicationInfoList_df, col_toparse)
                                                            

    # Saving the resulting dataframe to hive table
    applicationInfoList_df.repartition(200).write.format('parquet').partitionBy('part_date').mode('append').saveAsTable('TEST_BIGDATA_OWNER.DATASCORE_APPL_INFO')
    
    print('Execution time for applicationInfoList:', time.time() - start_time)
    print('')
    

def parse_externalImageInfoList(whole_data_hive):
    
    start_time = time.time()
    print('Currently parsing externalImageInfoList...')
    
    whole_data_hive.createOrReplaceTempView('test_bigdata_data_hive_temp')
    
    externalImageInfoList_df = spark.sql(
    '''
        select
               id_batch
              ,dtime_datascore
              ,part_date
              ,get_json_object(content, '$.externalImageInfoList[*].dateAdded') as dateAdded
              ,get_json_object(content, '$.externalImageInfoList[*].dateModified') as dateModified
              ,get_json_object(content, '$.externalImageInfoList[*].dateTaken') as dateTaken
              ,get_json_object(content, '$.externalImageInfoList[*].height') as height
              ,get_json_object(content, '$.externalImageInfoList[*].mimeType') as mimeType
              ,get_json_object(content, '$.externalImageInfoList[*].orientation') as orientation
              ,get_json_object(content, '$.externalImageInfoList[*].size') as size
              ,get_json_object(content, '$.externalImageInfoList[*].width') as width
        from test_bigdata_data_hive_temp
    '''
    
    )
    
    col_toparse = ['dateAdded', 'dateModified', 'dateTaken', 'height', 'mimeType', 'orientation', 'size', 'width']

    externalImageInfoList_df = cleanse_and_parse_func(externalImageInfoList_df, col_toparse)

    externalImageInfoList_df.repartition(200).write.format('parquet').partitionBy('part_date').mode('append').saveAsTable('TEST_BIGDATA_OWNER.DATASCORE_EXT_IMG_INFO')
    
    print('Execution time for externalImageInfoList:', time.time() - start_time)
    print('')


def parse_smsInfoList(whole_data_hive):
    
    start_time = time.time()
    print('Currently parsing smsInfoList...')
    
    whole_data_hive.createOrReplaceTempView('test_bigdata_data_hive_temp')
    
    smsInfoList_df = spark.sql(
        '''
            select
                   id_batch
                  ,dtime_datascore
                  ,part_date
                  ,get_json_object(content, '$.smsInfoList[*].creator') as creator
                  ,get_json_object(content, '$.smsInfoList[*].date') as date
                  ,get_json_object(content, '$.smsInfoList[*].dateSent') as dateSent
                  ,get_json_object(content, '$.smsInfoList[*].errorCode') as errorCode
                  ,get_json_object(content, '$.smsInfoList[*].hasSubject') as hasSubject
                  ,get_json_object(content, '$.smsInfoList[*].protocol') as protocol
                  ,get_json_object(content, '$.smsInfoList[*].read') as read
                  ,get_json_object(content, '$.smsInfoList[*].seen') as seen
                  ,get_json_object(content, '$.smsInfoList[*].serviceCenter') as serviceCenter
                  ,get_json_object(content, '$.smsInfoList[*].status') as status
                  ,get_json_object(content, '$.smsInfoList[*].subscriptionID') as subscriptionID
                  ,get_json_object(content, '$.smsInfoList[*].threadID') as threadID
                  ,get_json_object(content, '$.smsInfoList[*].type') as type
                  ,get_json_object(content, '$.smsInfoList[*].contactId') as contactId
            from test_bigdata_data_hive_temp
        '''
        )
        
    col_toparse = ['creator', 'date', 'dateSent', 'errorCode', 'hasSubject', 'protocol', 'read', 'seen', 'serviceCenter'
                    , 'status', 'subscriptionID', 'threadID', 'type', 'contactId']
                    
    smsInfoList_df = cleanse_and_parse_func(smsInfoList_df, col_toparse)
    
    smsInfoList_df.repartition(200).write.format('parquet').partitionBy('part_date').mode('append').saveAsTable('TEST_BIGDATA_OWNER.DATASCORE_SMS_INFO')
    
    print('Execution time for smsInfoList:', time.time() - start_time)
    print('')
    

def parse_downloadDataInfoList(whole_data_hive):
    
    start_time = time.time()
    print('Currently parsing downloadDataInfoList...')
    
    whole_data_hive.createOrReplaceTempView('test_bigdata_data_hive_temp')
    
    downloadDataInfoList_df = spark.sql(
        '''
            select
                   id_batch
                  ,dtime_datascore
                  ,part_date
                  ,get_json_object(content, '$.downloadDataInfoList[*].lastModified') as lastModified
                  ,get_json_object(content, '$.downloadDataInfoList[*].length') as length
                  ,get_json_object(content, '$.downloadDataInfoList[*].type') as type
            from test_bigdata_data_hive_temp
        '''
        )
        
    col_toparse = ['lastModified', 'length', 'type']
    
    downloadDataInfoList_df = cleanse_and_parse_func(downloadDataInfoList_df, col_toparse)
    
    downloadDataInfoList_df.repartition(200).write.format('parquet').partitionBy('part_date').mode('append').saveAsTable('TEST_BIGDATA_OWNER.DATASCORE_DOWNLOAD_INFO')
    
    print('Execution time for downloadDataInfoList:', time.time() - start_time)
    print('')
    
    
def parse_externalAudioInfoList(whole_data_hive):
    
    start_time = time.time()
    print('Currently parsing externalAudioInfoList...')
    
    whole_data_hive.createOrReplaceTempView('test_bigdata_data_hive_temp')
    
    externalAudioInfoList_df = spark.sql(
        '''
            select
                   id_batch
                  ,dtime_datascore
                  ,part_date
                  ,get_json_object(content, '$.externalAudioInfoList[*].dateAdded') as dateAdded
                  ,get_json_object(content, '$.externalAudioInfoList[*].dateModified') as dateModified
                  ,get_json_object(content, '$.externalAudioInfoList[*].duration') as duration
                  ,get_json_object(content, '$.externalAudioInfoList[*].mimeType') as mimeType
                  ,get_json_object(content, '$.externalAudioInfoList[*].music') as music
                  ,get_json_object(content, '$.externalAudioInfoList[*].year') as year
                  ,get_json_object(content, '$.externalAudioInfoList[*].isNotification') as isNotification
                  ,get_json_object(content, '$.externalAudioInfoList[*].isRingtone') as isRingtone
                  ,get_json_object(content, '$.externalAudioInfoList[*].isAlarm') as isAlarm
                  ,get_json_object(content, '$.externalAudioInfoList[*].album') as album
                  ,get_json_object(content, '$.externalAudioInfoList[*].albumID') as albumID
                  ,get_json_object(content, '$.externalAudioInfoList[*].albumKey') as albumKey
                  ,get_json_object(content, '$.externalAudioInfoList[*].artist') as artist
                  ,get_json_object(content, '$.externalAudioInfoList[*].artistID') as artistID
                  ,get_json_object(content, '$.externalAudioInfoList[*].artistKey') as artistKey
                  ,get_json_object(content, '$.externalAudioInfoList[*].bookmark') as bookmark
                  ,get_json_object(content, '$.externalAudioInfoList[*].composer') as composer
                  ,get_json_object(content, '$.externalAudioInfoList[*].isPodcast') as isPodcast
                  ,get_json_object(content, '$.externalAudioInfoList[*].titleKey') as titleKey
                  ,get_json_object(content, '$.externalAudioInfoList[*].track') as track
            from test_bigdata_data_hive_temp
        '''
        )
        
    col_toparse = ['dateAdded', 'dateModified', 'duration', 'mimeType', 'music', 'year', 'isNotification', 'isRingtone', 'isAlarm', 'album', 'albumID'
                    , 'albumKey', 'artist', 'artistID', 'artistKey', 'bookmark', 'composer', 'isPodcast', 'titleKey', 'track']
                    
    externalAudioInfoList_df = cleanse_and_parse_func(externalAudioInfoList_df, col_toparse)
    
    externalAudioInfoList_df.repartition(200).write.format('parquet').partitionBy('part_date').mode('append').saveAsTable('TEST_BIGDATA_OWNER.DATASCORE_EXT_AUD_INFO')
    
    print('Execution time for externalAudioInfoList:', time.time() - start_time)
    print('')
    

def parse_accountInfoList(whole_data_hive):
    
    start_time = time.time()
    print('Currently parsing accountInfoList...')
    
    whole_data_hive.createOrReplaceTempView('test_bigdata_data_hive_temp')
    
    accountInfoList_df = spark.sql(
        '''
            select
                   id_batch
                  ,dtime_datascore
                  ,part_date
                  ,get_json_object(content, '$.accountInfoList[*].accountType') as accountType
            from test_bigdata_data_hive_temp
        '''
        )
    
    col_toparse = ['accountType']
    
    accountInfoList_df = cleanse_and_parse_func(accountInfoList_df, col_toparse)
    
    accountInfoList_df.repartition(200).write.format('parquet').partitionBy('part_date').mode('append').saveAsTable('TEST_BIGDATA_OWNER.DATASCORE_ACCT_INFO')
    
    print('Execution time for accountInfoList:', time.time() - start_time)
    print('')
    

def parse_bluetoothInfo(whole_data_hive):
    
    start_time = time.time()
    print('Currently parsing bluetoothInfo...')
    
    whole_data_hive.createOrReplaceTempView('test_bigdata_data_hive_temp')
    
    bluetoothInfo_df = spark.sql(
        '''
            select
                   id_batch
                  ,dtime_datascore
                  ,part_date
                  ,get_json_object(content, '$.bluetoothInfo[*].address') as address
                  ,get_json_object(content, '$.bluetoothInfo[*].bondState') as bondState
                  ,get_json_object(content, '$.bluetoothInfo[*].deviceClass') as deviceClass
                  ,get_json_object(content, '$.bluetoothInfo[*].majorDeviceClass') as majorDeviceClass
                  ,get_json_object(content, '$.bluetoothInfo[*].name') as name
                  ,get_json_object(content, '$.bluetoothInfo[*].type') as type
                  ,get_json_object(content, '$.bluetoothInfo[*].UUIDS') as UUIDS
            from test_bigdata_data_hive_temp
        '''
        )
        
    col_toparse = ['address', 'bondState', 'deviceClass', 'majorDeviceClass', 'name', 'type', 'UUIDS']
    
    bluetoothInfo_df = cleanse_and_parse_func(bluetoothInfo_df, col_toparse)
    
    bluetoothInfo_df.repartition(200).write.format('parquet').partitionBy('part_date').mode('append').saveAsTable('TEST_BIGDATA_OWNER.DATASCORE_BLUETOOTH_INFO')
    
    print('Execution time for bluetoothInfo:', time.time() - start_time)
    print('')
    

def parse_calendarEventsInfoList(whole_data_hive):
    
    start_time = time.time()
    print('Currently parsing calendarEventsInfoList...')
    
    whole_data_hive.createOrReplaceTempView('test_bigdata_data_hive_temp')
    
    calendarEventsInfoList_df = spark.sql(
        '''
            select
                   id_batch
                  ,dtime_datascore
                  ,part_date
                  ,get_json_object(content, '$.calendarEventsInfoList[*].guestsCanInviteOthers') as guestsCanInviteOthers
                  ,get_json_object(content, '$.calendarEventsInfoList[*].guestsCanModify') as guestsCanModify
                  ,get_json_object(content, '$.calendarEventsInfoList[*].guestsCanSeeGuests') as guestsCanSeeGuests
                  ,get_json_object(content, '$.calendarEventsInfoList[*].hasAlarm') as hasAlarm
                  ,get_json_object(content, '$.calendarEventsInfoList[*].hasAttendeeData') as hasAttendeeData
                  ,get_json_object(content, '$.calendarEventsInfoList[*].hasExtendedProperties') as hasExtendedProperties
                  ,get_json_object(content, '$.calendarEventsInfoList[*].isOrganizer') as isOrganizer
                  ,get_json_object(content, '$.calendarEventsInfoList[*].lastDate') as lastDate
                  ,get_json_object(content, '$.calendarEventsInfoList[*].rrule') as rrule
                  ,get_json_object(content, '$.calendarEventsInfoList[*].status') as status
                  ,get_json_object(content, '$.calendarEventsInfoList[*].rdate') as rdate
            from test_bigdata_data_hive_temp
        '''
        )
    
    col_toparse = ['guestsCanInviteOthers', 'guestsCanModify', 'guestsCanSeeGuests', 'hasAlarm', 'hasAttendeeData', 'hasExtendedProperties'
                    , 'isOrganizer', 'lastDate', 'rrule', 'status', 'rdate']
                    
    calendarEventsInfoList_df = cleanse_and_parse_func(calendarEventsInfoList_df, col_toparse)
    
    calendarEventsInfoList_df.repartition(200).write.format('parquet').partitionBy('part_date').mode('append').saveAsTable('TEST_BIGDATA_OWNER.DATASCORE_CALENDAR_EVENTS_INFO')
    
    print('Execution time for calendarEventsInfoList:', time.time() - start_time)
    print('')
    
    
def parse_calendarInfoList(whole_data_hive):
    
    start_time = time.time()
    print('Currently parsing calendarInfoList...')
    
    whole_data_hive.createOrReplaceTempView('test_bigdata_data_hive_temp')
    
    calendarInfoList_df = spark.sql(
        '''
            select
                   id_batch
                  ,dtime_datascore
                  ,part_date
                  ,get_json_object(content, '$.calendarInfoList[*].accountType') as accountType
                  ,get_json_object(content, '$.calendarInfoList[*].allowedAttendeeTypes') as allowedAttendeeTypes
                  ,get_json_object(content, '$.calendarInfoList[*].allowedAvailability') as allowedAvailability
                  ,get_json_object(content, '$.calendarInfoList[*].allowedReminders') as allowedReminders
                  ,get_json_object(content, '$.calendarInfoList[*].calendarAccessLevel') as calendarAccessLevel
                  ,get_json_object(content, '$.calendarInfoList[*].calendarTimeZone') as calendarTimeZone
                  ,get_json_object(content, '$.calendarInfoList[*].id') as id
                  ,get_json_object(content, '$.calendarInfoList[*].syncEvents') as syncEvents
                  ,get_json_object(content, '$.calendarInfoList[*].visible') as visible
            from test_bigdata_data_hive_temp
        '''
        )
        
    col_toparse = ['accountType', 'allowedAttendeeTypes', 'allowedAvailability', 'allowedReminders', 'calendarAccessLevel'
                    , 'calendarTimeZone', 'id', 'syncEvents', 'visible']
                    
    calendarInfoList_df = cleanse_and_parse_func(calendarInfoList_df, col_toparse)
    
    calendarInfoList_df.repartition(200).write.format('parquet').partitionBy('part_date').mode('append').saveAsTable('TEST_BIGDATA_OWNER.DATASCORE_CALENDAR_INFO')
    
    print('Execution time for calendarInfoList:', time.time() - start_time)
    print('')
    
    
def parse_calendarRemindersInfoList(whole_data_hive):
    
    start_time = time.time()
    print('Currently parsing calendarRemindersInfoList...')
    
    whole_data_hive.createOrReplaceTempView('test_bigdata_data_hive_temp')
    
    calendarRemindersInfoList_df = spark.sql(
        '''
            select
                   id_batch
                  ,dtime_datascore
                  ,part_date
                  ,get_json_object(content, '$.calendarRemindersInfoList[*].eventId') as eventId
                  ,get_json_object(content, '$.calendarRemindersInfoList[*].method') as method
                  ,get_json_object(content, '$.calendarRemindersInfoList[*].minutes') as minutes
            from test_bigdata_data_hive_temp
        '''
        )
    
    col_toparse = ['eventId', 'method', 'minutes']
    
    calendarRemindersInfoList_df = cleanse_and_parse_func(calendarRemindersInfoList_df, col_toparse)
    
    calendarRemindersInfoList_df.repartition(200).write.format('parquet').partitionBy('part_date').mode('append').saveAsTable('TEST_BIGDATA_OWNER.DATASCORE_CALENDAR_REMINDER_INFO')
    
    print('Execution time for calendarRemindersInfoList:', time.time() - start_time)
    print('')
    
    
def parse_calendarAttendeeInfoList(whole_data_hive):
    
    start_time = time.time()
    print('Currently parsing calendarAttendeeInfoList...')
    
    whole_data_hive.createOrReplaceTempView('test_bigdata_data_hive_temp')
    
    calendarAttendeeInfoList_df = spark.sql(
        '''
            select
                   id_batch
                  ,dtime_datascore
                  ,part_date
                  ,get_json_object(content, '$.calendarAttendeeInfoList[*].eventId') as eventId
                  ,get_json_object(content, '$.calendarAttendeeInfoList[*].attendeeType') as attendeeType
                  ,get_json_object(content, '$.calendarAttendeeInfoList[*].attendeeStatus') as attendeeStatus
            from test_bigdata_data_hive_temp
        '''
        )
        
    col_toparse = ['eventId', 'attendeeType', 'attendeeStatus']
    
    calendarAttendeeInfoList_df = cleanse_and_parse_func(calendarAttendeeInfoList_df, col_toparse)
    
    calendarAttendeeInfoList_df.repartition(200).write.format('parquet').partitionBy('part_date').mode('append').saveAsTable('TEST_BIGDATA_OWNER.DATASCORE_CALENDAR_ATTENDEE_INFO')
    
    print('Execution time for calendarAttendeeInfoList:', time.time() - start_time)
    print('')

def parse_contactInfoList(whole_data_hive):
    
    start_time = time.time()
    print('Currently parsing contactInfoList...')
    
    whole_data_hive.createOrReplaceTempView('test_bigdata_data_hive_temp')
    
    contactInfoList_df = spark.sql(
        '''
            select
                   id_batch
                  ,dtime_datascore
                  ,part_date
                  ,get_json_object(content, '$.contactInfoList[*].contactLastUpdatedTimestamp') as contactLastUpdatedTimestamp
                  ,get_json_object(content, '$.contactInfoList[*].hasContactStatus') as hasContactStatus
                  ,get_json_object(content, '$.contactInfoList[*].hasPhoneNumber') as hasPhoneNumber
                  ,get_json_object(content, '$.contactInfoList[*].hasPhotoId') as hasPhotoId
                  ,get_json_object(content, '$.contactInfoList[*].id') as id
                  ,get_json_object(content, '$.contactInfoList[*].inDefaultDirectory') as inDefaultDirectory
                  ,get_json_object(content, '$.contactInfoList[*].inVisibleGroup') as inVisibleGroup
                  ,get_json_object(content, '$.contactInfoList[*].isCustomRingtone') as isCustomRingtone
                  ,get_json_object(content, '$.contactInfoList[*].isUserProfile') as isUserProfile
                  ,get_json_object(content, '$.contactInfoList[*].lastTimeContacted') as lastTimeContacted
                  ,get_json_object(content, '$.contactInfoList[*].photoFileID') as photoFileID
                  ,get_json_object(content, '$.contactInfoList[*].sendToVoicemail') as sendToVoicemail
                  ,get_json_object(content, '$.contactInfoList[*].starred') as starred
                  ,get_json_object(content, '$.contactInfoList[*].timesContacted') as timesContacted
                  ,get_json_object(content, '$.contactInfoList[*].contactStatusTimestamp') as contactStatusTimestamp
            from test_bigdata_data_hive_temp
        '''
        )
    
    col_toparse = ['contactLastUpdatedTimestamp', 'hasContactStatus', 'hasPhoneNumber', 'hasPhotoId', 'id', 'inDefaultDirectory', 'inVisibleGroup', 'isCustomRingtone'
                    , 'isUserProfile', 'lastTimeContacted', 'photoFileID', 'sendToVoicemail', 'starred', 'timesContacted', 'contactStatusTimestamp']
                    
    contactInfoList_df = cleanse_and_parse_func(contactInfoList_df, col_toparse)
    
    contactInfoList_df.repartition(200).write.format('parquet').partitionBy('part_date').mode('append').saveAsTable('TEST_BIGDATA_OWNER.DATASCORE_CONTACT_INFO')
    
    print('Execution time for contactInfoList:', time.time() - start_time)
    print('')


def parse_errors(whole_data_hive):
    
    start_time = time.time()
    print('Currently parsing errors...')
    
    whole_data_hive.createOrReplaceTempView('test_bigdata_data_hive_temp')
    
    errors_df = spark.sql(
        '''
            select
                   id_batch
                  ,dtime_datascore
                  ,part_date
                  ,get_json_object(content, '$.errors[*].error') as error
                  ,get_json_object(content, '$.errors[*].fieldName') as fieldName
            from test_bigdata_data_hive_temp
        '''
        )
        
    col_toparse = ['error', 'fieldName']
    
    errors_df = cleanse_and_parse_func(errors_df, col_toparse)
    
    errors_df.repartition(200).write.format('parquet').partitionBy('part_date').mode('append').saveAsTable('TEST_BIGDATA_OWNER.DATASCORE_ERRORS')
    
    print('Execution time for errors:', time.time() - start_time)
    print('')
    
    
def parse_fileInfoList(whole_data_hive):
    
    start_time = time.time()
    print('Currently parsing fileInfoList...')
    
    whole_data_hive.createOrReplaceTempView('test_bigdata_data_hive_temp')
    
    fileInfoList_df = spark.sql(
        '''
            select
                   id_batch
                  ,dtime_datascore
                  ,part_date
                  ,get_json_object(content, '$.fileInfoList[*].location') as location
                  ,get_json_object(content, '$.fileInfoList[*].size') as size
                  ,get_json_object(content, '$.fileInfoList[*].type') as type
            from test_bigdata_data_hive_temp
        '''
        )
        
    col_toparse = ['location', 'size', 'type']
    
    fileInfoList_df = cleanse_and_parse_func(fileInfoList_df, col_toparse)
    
    fileInfoList_df.repartition(200).write.format('parquet').partitionBy('part_date').mode('append').saveAsTable('TEST_BIGDATA_OWNER.DATASCORE_FILE_INFO')
    
    print('Execution time for fileInfoList:', time.time() - start_time)
    print('')
    
    
def parse_gmailInfoList(whole_data_hive):
    
    start_time = time.time()
    print('Currently parsing gmailInfoList...')
    
    whole_data_hive.createOrReplaceTempView('test_bigdata_data_hive_temp')
    
    gmailInfoList_df = spark.sql(
        '''
            select
                   id_batch
                  ,dtime_datascore
                  ,part_date
                  ,get_json_object(content, '$.gmailInfoList[*].id') as id
                  ,get_json_object(content, '$.gmailInfoList[*].name') as name
                  ,get_json_object(content, '$.gmailInfoList[*].numConversations') as numConversations
                  ,get_json_object(content, '$.gmailInfoList[*].numUnreadConversations') as numUnreadConversations
            from daa_hive_temp
        '''
        )
    
    col_toparse = ['id', 'name', 'numConversations', 'numUnreadConversations']
    
    gmailInfoList_df = cleanse_and_parse_func(gmailInfoList_df, col_toparse)
    
    gmailInfoList_df.repartition(200).write.format('parquet').partitionBy('part_date').mode('append').saveAsTable('TEST_BIGDATA_OWNER.DATASCORE_GMAIL_INFO')
    
    print('Execution time for gmailInfoList:', time.time() - start_time)
    print('')
    
    
def parse_internalAudioInfoList(whole_data_hive):
    
    start_time = time.time()
    print('Currently parsing internalAudioInfoList...')
    
    whole_data_hive.createOrReplaceTempView('test_bigdata_data_hive_temp')
    
    internalAudioInfoList_df = spark.sql(
        '''
            select
                   id_batch
                  ,dtime_datascore
                  ,part_date
                  ,get_json_object(content, '$.internalAudioInfoList[*].album') as album
                  ,get_json_object(content, '$.internalAudioInfoList[*].albumID') as albumID
                  ,get_json_object(content, '$.internalAudioInfoList[*].albumKey') as albumKey
                  ,get_json_object(content, '$.internalAudioInfoList[*].artist') as artist
                  ,get_json_object(content, '$.internalAudioInfoList[*].artistID') as artistID
                  ,get_json_object(content, '$.internalAudioInfoList[*].artistKey') as artistKey
                  ,get_json_object(content, '$.internalAudioInfoList[*].dateAdded') as dateAdded
                  ,get_json_object(content, '$.internalAudioInfoList[*].dateModified') as dateModified
                  ,get_json_object(content, '$.internalAudioInfoList[*].duration') as duration
                  ,get_json_object(content, '$.internalAudioInfoList[*].isAlarm') as isAlarm
                  ,get_json_object(content, '$.internalAudioInfoList[*].isNotification') as isNotification
                  ,get_json_object(content, '$.internalAudioInfoList[*].isPodcast') as isPodcast
                  ,get_json_object(content, '$.internalAudioInfoList[*].isRingtone') as isRingtone
                  ,get_json_object(content, '$.internalAudioInfoList[*].mimeType') as mimeType
                  ,get_json_object(content, '$.internalAudioInfoList[*].music') as music
                  ,get_json_object(content, '$.internalAudioInfoList[*].titleKey') as titleKey
                  ,get_json_object(content, '$.internalAudioInfoList[*].track') as track
                  ,get_json_object(content, '$.internalAudioInfoList[*].year') as year
                  ,get_json_object(content, '$.internalAudioInfoList[*].composer') as composer
            from test_bigdata_data_hive_temp
        '''
        )
    
    col_toparse = ['album', 'albumID', 'albumKey', 'artist', 'artistID', 'artistKey', 'dateAdded', 'dateModified', 'duration', 'isAlarm', 'isNotification'
                    , 'isPodcast', 'isRingtone', 'mimeType', 'music', 'titleKey', 'track', 'year', 'composer']
                    
    internalAudioInfoList_df = cleanse_and_parse_func(internalAudioInfoList_df, col_toparse)
    
    internalAudioInfoList_df.repartition(200).write.format('parquet').partitionBy('part_date').mode('append').saveAsTable('TEST_BIGDATA_OWNER.DATASCORE_INT_AUD_INFO')
    
    print('Execution time for internalAudioInfoList:', time.time() - start_time)
    print('')


def parse_internalImageInfoList(whole_data_hive):
    
    start_time = time.time()
    print('Currently parsing internalImageInfoList...')
    
    whole_data_hive.createOrReplaceTempView('test_bigdata_data_hive_temp')
    
    internalImageInfoList_df = spark.sql(
        '''
            select
                   id_batch
                  ,dtime_datascore
                  ,part_date
                  ,get_json_object(content, '$.internalImageInfoList[*].dateAdded') as dateAdded
                  ,get_json_object(content, '$.internalImageInfoList[*].dateModified') as dateModified
                  ,get_json_object(content, '$.internalImageInfoList[*].dateTaken') as dateTaken
                  ,get_json_object(content, '$.internalImageInfoList[*].height') as height
                  ,get_json_object(content, '$.internalImageInfoList[*].mimeType') as mimeType
                  ,get_json_object(content, '$.internalImageInfoList[*].size') as size
                  ,get_json_object(content, '$.internalImageInfoList[*].width') as width
                  ,get_json_object(content, '$.internalImageInfoList[*].orientation') as orientation
            from test_bigdata_data_hive_temp
        '''
        )
        
    col_toparse = ['dateAdded', 'dateModified', 'dateTaken', 'height', 'mimeType', 'size', 'width', 'orientation']
    
    internalImageInfoList_df = cleanse_and_parse_func(internalImageInfoList_df, col_toparse)
    
    internalImageInfoList_df.repartition(200).write.format('parquet').partitionBy('part_date').mode('append').saveAsTable('TEST_BIGDATA_OWNER.DATASCORE_INT_IMG_INFO')
    
    print('Execution time for internalImageInfoList:', time.time() - start_time)
    print('')
    
    
def parse_internalVideoInfoList(whole_data_hive):
    
    start_time = time.time()
    print('Currently parsing internalVideoInfoList...')
    
    whole_data_hive.createOrReplaceTempView('test_bigdata_data_hive_temp')
    
    internalVideoInfoList_df = spark.sql(
        '''
            select
                   id_batch
                  ,dtime_datascore
                  ,part_date
                  ,get_json_object(content, '$.internalVideoInfoList[*].dateAdded') as dateAdded
                  ,get_json_object(content, '$.internalVideoInfoList[*].dateModified') as dateModified
                  ,get_json_object(content, '$.internalVideoInfoList[*].dateTaken') as dateTaken
                  ,get_json_object(content, '$.internalVideoInfoList[*].duration') as duration
                  ,get_json_object(content, '$.internalVideoInfoList[*].hasTags') as hasTags
                  ,get_json_object(content, '$.internalVideoInfoList[*].iPrivate') as iPrivate
                  ,get_json_object(content, '$.internalVideoInfoList[*].mimeType') as mimeType
                  ,get_json_object(content, '$.internalVideoInfoList[*].resolution') as resolution
                  ,get_json_object(content, '$.internalVideoInfoList[*].size') as size
            from test_bigdata_data_hive_temp
        '''
        )
        
    col_toparse = ['dateAdded', 'dateModified', 'dateTaken', 'duration', 'hasTags', 'iPrivate', 'mimeType', 'resolution', 'size']
    
    internalVideoInfoList_df = cleanse_and_parse_func(internalVideoInfoList_df, col_toparse)
    
    internalVideoInfoList_df.repartition(200).write.format('parquet').partitionBy('part_date').mode('append').saveAsTable('TEST_BIGDATA_OWNER.DATASCORE_INT_VID_INFO')
    
    print('Execution time for internalVideoInfoList:', time.time() - start_time)
    print('')
    

def parse_sensorInfoList(whole_data_hive):
    
    start_time = time.time()
    print('Currently parsing sensorInfoList...')
    
    whole_data_hive.createOrReplaceTempView('test_bigdata_data_hive_temp')
    
    sensorInfoList_df = spark.sql(
        '''
            select
                   id_batch
                  ,dtime_datascore
                  ,part_date
                  ,get_json_object(content, '$.sensorInfoList[*].additionalInfoSupported') as additionalInfoSupported
                  ,get_json_object(content, '$.sensorInfoList[*].dynamicSensor') as dynamicSensor
                  ,get_json_object(content, '$.sensorInfoList[*].fifoMaxEventCount') as fifoMaxEventCount
                  ,get_json_object(content, '$.sensorInfoList[*].fifoReservedEventCount') as fifoReservedEventCount
                  ,get_json_object(content, '$.sensorInfoList[*].highestDirectReportRateLevel') as highestDirectReportRateLevel
                  ,get_json_object(content, '$.sensorInfoList[*].id') as id
                  ,get_json_object(content, '$.sensorInfoList[*].isWakeUpSensor') as isWakeUpSensor
                  ,get_json_object(content, '$.sensorInfoList[*].maxDelay') as maxDelay
                  ,get_json_object(content, '$.sensorInfoList[*].maximumRange') as maximumRange
                  ,get_json_object(content, '$.sensorInfoList[*].minDelay') as minDelay
                  ,get_json_object(content, '$.sensorInfoList[*].name') as name
                  ,get_json_object(content, '$.sensorInfoList[*].power') as power
                  ,get_json_object(content, '$.sensorInfoList[*].reportingMode') as reportingMode
                  ,get_json_object(content, '$.sensorInfoList[*].resolution') as resolution
                  ,get_json_object(content, '$.sensorInfoList[*].stringType') as stringType
                  ,get_json_object(content, '$.sensorInfoList[*].type') as type
                  ,get_json_object(content, '$.sensorInfoList[*].vendor') as vendor
                  ,get_json_object(content, '$.sensorInfoList[*].version') as version
            from test_bigdata_data_hive_temp
        '''
        )
        
    col_toparse = ['additionalInfoSupported', 'dynamicSensor', 'fifoMaxEventCount', 'fifoReservedEventCount', 'highestDirectReportRateLevel', 'id', 'isWakeUpSensor'
                    , 'maxDelay', 'maximumRange', 'minDelay', 'name', 'power', 'reportingMode', 'resolution', 'stringType', 'type', 'vendor', 'version']
                    
    sensorInfoList_df = cleanse_and_parse_func(sensorInfoList_df, col_toparse)
    
    sensorInfoList_df.repartition(200).write.format('parquet').partitionBy('part_date').mode('append').saveAsTable('TEST_BIGDATA_OWNER.DATASCORE_SENSOR_INFO')
    
    print('Execution time for sensorInfoList:', time.time() - start_time)
    print('')
    

def parse_externalVideoInfoList(whole_data_hive):
    
    start_time = time.time()
    print('Currently parsing externalVideoInfoList...')
    
    whole_data_hive.createOrReplaceTempView('test_bigdata_data_hive_temp')
    
    externalVideoInfoList_df = spark.sql(
        '''
            select
                   id_batch
                  ,dtime_datascore
                  ,part_date
                  ,get_json_object(content, '$.externalVideoInfoList[*].dateAdded') as dateAdded
                  ,get_json_object(content, '$.externalVideoInfoList[*].dateModified') as dateModified
                  ,get_json_object(content, '$.externalVideoInfoList[*].dateTaken') as dateTaken
                  ,get_json_object(content, '$.externalVideoInfoList[*].duration') as duration
                  ,get_json_object(content, '$.externalVideoInfoList[*].hasTags') as hasTags
                  ,get_json_object(content, '$.externalVideoInfoList[*].iPrivate') as iPrivate
                  ,get_json_object(content, '$.externalVideoInfoList[*].mimeType') as mimeType
                  ,get_json_object(content, '$.externalVideoInfoList[*].resolution') as resolution
                  ,get_json_object(content, '$.externalVideoInfoList[*].size') as size
            from test_bigdata_data_hive_temp
        '''
        )
    
    col_toparse = ['dateAdded', 'dateModified', 'dateTaken', 'duration', 'hasTags', 'iPrivate', 'mimeType', 'resolution', 'size']
    
    externalVideoInfoList_df = cleanse_and_parse_func(externalVideoInfoList_df, col_toparse)
    
    externalVideoInfoList_df.repartition(200).write.format('parquet').partitionBy('part_date').mode('append').saveAsTable('TEST_BIGDATA_OWNER.DATASCORE_EXT_VID_INFO')
    
    print('Execution time for externalVideoInfoList:', time.time() - start_time)
    print('')
    

def parse_contactGroupInfoList(whole_data_hive):
    
    start_time = time.time()
    print('Currently parsing contactGroupInfoList...')
    
    whole_data_hive.createOrReplaceTempView('test_bigdata_data_hive_temp')
    
    contactGroupInfoList_df = spark.sql(
        '''
            select
                   id_batch
                  ,dtime_datascore
                  ,part_date
                  ,get_json_object(content, '$.contactGroupInfoList[*].accountType') as accountType
                  ,get_json_object(content, '$.contactGroupInfoList[*].autoAdd') as autoAdd
                  ,get_json_object(content, '$.contactGroupInfoList[*].deleted') as deleted
                  ,get_json_object(content, '$.contactGroupInfoList[*].favorites') as favorites
                  ,get_json_object(content, '$.contactGroupInfoList[*].groupIsReadOnly') as groupIsReadOnly
                  ,get_json_object(content, '$.contactGroupInfoList[*].groupVisible') as groupVisible
                  ,get_json_object(content, '$.contactGroupInfoList[*].hasNotes') as hasNotes
                  ,get_json_object(content, '$.contactGroupInfoList[*].id') as id
                  ,get_json_object(content, '$.contactGroupInfoList[*].shouldSync') as shouldSync
            from test_bigdata_data_hive_temp
        '''
        )
        
    col_toparse = ['accountType', 'autoAdd', 'deleted', 'favorites', 'groupIsReadOnly', 'groupVisible', 'hasNotes', 'id', 'shouldSync']    
    
    contactGroupInfoList_df = cleanse_and_parse_func(contactGroupInfoList_df, col_toparse)
    
    contactGroupInfoList_df.repartition(200).write.format('parquet').partitionBy('part_date').mode('append').saveAsTable('TEST_BIGDATA_OWNER.DATASCORE_CONTACT_GROUP_INFO')
    
    print('Execution time for contactGroupInfoList:', time.time() - start_time)
    print('')
    

def parse_batteryInfo(whole_data_hive):
    
    start_time = time.time()
    print('Currently parsing batteryInfo...')
    
    whole_data_hive.createOrReplaceTempView('test_bigdata_data_hive_temp')
    
    batteryInfo_df = spark.sql(
        '''
            select
                   id_batch
                  ,dtime_datascore
                  ,part_date
                  ,get_json_object(content, '$.batteryInfo.batteryPct') as batteryPct
                  ,get_json_object(content, '$.batteryInfo.chargePlug') as chargePlug
                  ,get_json_object(content, '$.batteryInfo.health') as health
                  ,get_json_object(content, '$.batteryInfo.iconSmall') as iconSmall
                  ,get_json_object(content, '$.batteryInfo.isCharging') as isCharging
                  ,get_json_object(content, '$.batteryInfo.level') as level
                  ,get_json_object(content, '$.batteryInfo.present') as present
                  ,get_json_object(content, '$.batteryInfo.scale') as scale
                  ,get_json_object(content, '$.batteryInfo.status') as status
                  ,get_json_object(content, '$.batteryInfo.technology') as technology
                  ,get_json_object(content, '$.batteryInfo.temperatur') as temperatur
                  ,get_json_object(content, '$.batteryInfo.voltage') as voltage
            from test_bigdata_data_hive_temp
        '''
        )
        
    col_toparse = ['batteryPct', 'chargePlug', 'health', 'iconSmall', 'isCharging', 'level', 'present', 'scale', 'status', 'technology', 'temperatur', 'voltage']
    
    batteryInfo_df = batteryInfo_df.select(col_toparse + ['id_batch', 'dtime_datascore', 'part_date'])
        
    batteryInfo_df.repartition(200).write.format('parquet').partitionBy('part_date').mode('append').saveAsTable('TEST_BIGDATA_OWNER.DATASCORE_BATTERY_INFO')
    
    print('Execution time for batteryInfo:', time.time() - start_time)
    print('')
    
    
def parse_generalInfo(whole_data_hive):
    
    start_time = time.time()
    print('Currently parsing generalInfo...')
    
    whole_data_hive.createOrReplaceTempView('test_bigdata_data_hive_temp')
    
    generalInfo_df = spark.sql(
        '''
            select
                   id_batch
                  ,dtime_datascore
                  ,part_date
                  ,get_json_object(content, '$.generalInfo.androidId') as androidId
                  ,get_json_object(content, '$.generalInfo.imei') as imei
                  ,get_json_object(content, '$.generalInfo.keyboard') as keyboard
                  ,get_json_object(content, '$.generalInfo.localeDisplayLanguage') as localeDisplayLanguage
                  ,get_json_object(content, '$.generalInfo.mcc') as mcc
                  ,get_json_object(content, '$.generalInfo.meid') as meid
                  ,get_json_object(content, '$.generalInfo.mnc') as mnc
                  ,get_json_object(content, '$.generalInfo.networkOperator') as networkOperator
                  ,get_json_object(content, '$.generalInfo.networkOperatorName') as networkOperatorName
                  ,get_json_object(content, '$.generalInfo.phoneNumber') as phoneNumber
                  ,get_json_object(content, '$.generalInfo.phoneType') as phoneType
                  ,get_json_object(content, '$.generalInfo.simCountryIso') as simCountryIso
                  ,get_json_object(content, '$.generalInfo.simSerialNumber') as simSerialNumber
                  ,get_json_object(content, '$.generalInfo.timeZoneId') as timeZoneId
                  ,get_json_object(content, '$.generalInfo.deviceId') as deviceId
                  ,get_json_object(content, '$.generalInfo.localeIso3Country') as localeIso3Country
                  ,get_json_object(content, '$.generalInfo.localeIso3Language') as localeIso3Language
            from test_bigdata_data_hive_temp
        '''
        )
    
    col_toparse = ['androidId', 'imei', 'keyboard', 'localeDisplayLanguage', 'mcc', 'meid', 'mnc', 'networkOperator', 'networkOperatorName'
                    , 'phoneNumber', 'phoneType', 'simCountryIso', 'simSerialNumber', 'timeZoneId', 'deviceId', 'localeIso3Country', 'localeIso3Language']
                    
    
    for col in col_toparse:
        generalInfo_df = generalInfo_df.withColumn(col, regexp_replace(col,'"',""))
        
    generalInfo_df = generalInfo_df.select(col_toparse + ['id_batch', 'dtime_datascore', 'part_date'])
    
    generalInfo_df.repartition(200).write.format('parquet').partitionBy('part_date').mode('append').saveAsTable('TEST_BIGDATA_OWNER.DATASCORE_GENERAL_INFO')
    
    print('Execution time for generalInfo:', time.time() - start_time)
    print('')


def parse_libraryInfo(whole_data_hive):
    
    start_time = time.time()
    print('Currently parsing libraryInfo...')
    
    whole_data_hive.createOrReplaceTempView('test_bigdata_data_hive_temp')
    
    libraryInfo_df = spark.sql(
        '''
            select
                   id_batch
                  ,dtime_datascore
                  ,part_date
                  ,get_json_object(content, '$.libraryInfo.endCollectTime') as endCollectTime
                  ,get_json_object(content, '$.libraryInfo.libraryVersion') as libraryVersion
                  ,get_json_object(content, '$.libraryInfo.startCollectTime') as startCollectTime
            from test_bigdata_data_hive_temp
        '''
        )
        
    col_toparse = ['endCollectTime', 'libraryVersion', 'startCollectTime']
    
    for col in col_toparse:
        libraryInfo_df = libraryInfo_df.withColumn(col, regexp_replace(col,'"',""))
        
    libraryInfo_df = libraryInfo_df.select(col_toparse + ['id_batch', 'dtime_datascore', 'part_date'])
    
    libraryInfo_df.repartition(200).write.format('parquet').partitionBy('part_date').mode('append').saveAsTable('TEST_BIGDATA_OWNER.DATASCORE_LIBRARY_INFO')
    
    print('Execution time for libraryInfo:', time.time() - start_time)
    print('')
    

def parse_wifiInfo(whole_data_hive):
    
    start_time = time.time()
    print('Currently parsing wifiInfo...')
    
    whole_data_hive.createOrReplaceTempView('test_bigdata_data_hive_temp')
    
    wifiInfo_df = spark.sql(
        '''
            select
                   id_batch
                  ,dtime_datascore
                  ,part_date
                  ,get_json_object(content, '$.wifiInfo.bssid') as bssid
                  ,get_json_object(content, '$.wifiInfo.frequency') as frequency
                  ,get_json_object(content, '$.wifiInfo.ipAddress') as ipAddress
                  ,get_json_object(content, '$.wifiInfo.macAddress') as macAddress
                  ,get_json_object(content, '$.wifiInfo.wifiConfigList') as wifiConfigList
            from test_bigdata_data_hive_temp
        '''
        )
        
    col_toparse = ['bssid', 'frequency', 'ipAddress', 'macAddress', 'wifiConfigList']
    
    for col in col_toparse:
        wifiInfo_df = wifiInfo_df.withColumn(col, regexp_replace(col,'"',""))
        
    wifiInfo_df = wifiInfo_df.select(col_toparse + ['id_batch', 'dtime_datascore', 'part_date'])
    
    wifiInfo_df.repartition(200).write.format('parquet').partitionBy('part_date').mode('append').saveAsTable('TEST_BIGDATA_OWNER.DATASCORE_WIFI_INFO')
    
    print('Execution time for wifiInfo:', time.time() - start_time)
    print('')
    

def parse_locationInfo(whole_data_hive):
    
    start_time = time.time()
    print('Currently parsing locationInfo...')
    
    whole_data_hive.createOrReplaceTempView('test_bigdata_data_hive_temp')
    
    locationInfo_df = spark.sql(
        '''
            select
                   id_batch
                  ,dtime_datascore
                  ,part_date
                  ,get_json_object(content, '$.locationInfo.address') as address
            from test_bigdata_data_hive_temp
        '''
        )
        
    col_toparse = ['address']
    
    for col in col_toparse:
        locationInfo_df = locationInfo_df.withColumn(col, regexp_replace(col,'"',""))
        
    locationInfo_df = locationInfo_df.select(col_toparse + ['id_batch', 'dtime_datascore', 'part_date'])
    
    locationInfo_df.repartition(200).write.format('parquet').partitionBy('part_date').mode('append').saveAsTable('TEST_BIGDATA_OWNER.DATASCORE_LOCATION_INFO')
    
    print('Execution time for locationInfo:', time.time() - start_time)
    print('')
    

def parse_hardwareInfo(whole_data_hive):
    
    start_time = time.time()
    print('Currently parsing hardwareInfo...')
    
    whole_data_hive.createOrReplaceTempView('test_bigdata_data_hive_temp')
    
    hardwareInfo_df = spark.sql(
        '''
            select
                   id_batch
                  ,dtime_datascore
                  ,part_date
                  ,get_json_object(content, '$.hardwareInfo.bootloader') as bootloader
                  ,get_json_object(content, '$.hardwareInfo.brand') as brand
                  ,get_json_object(content, '$.hardwareInfo.display') as display
                  ,get_json_object(content, '$.hardwareInfo.externalStorageFree') as externalStorageFree
                  ,get_json_object(content, '$.hardwareInfo.externalStorageTotal') as externalStorageTotal
                  ,get_json_object(content, '$.hardwareInfo.fingerprint') as fingerprint
                  ,get_json_object(content, '$.hardwareInfo.hardware') as hardware
                  ,get_json_object(content, '$.hardwareInfo.host') as host
                  ,get_json_object(content, '$.hardwareInfo.id') as id
                  ,get_json_object(content, '$.hardwareInfo.mainStorageFree') as mainStorageFree
                  ,get_json_object(content, '$.hardwareInfo.mainStorageTotal') as mainStorageTotal
                  ,get_json_object(content, '$.hardwareInfo.manufacturer') as manufacturer
                  ,get_json_object(content, '$.hardwareInfo.model') as model
                  ,get_json_object(content, '$.hardwareInfo.product') as product
                  ,get_json_object(content, '$.hardwareInfo.ramTotalSize') as ramTotalSize
                  ,get_json_object(content, '$.hardwareInfo.release') as release
                  ,get_json_object(content, '$.hardwareInfo.serial') as serial
                  ,get_json_object(content, '$.hardwareInfo.supported32BitAbis') as supported32BitAbis
                  ,get_json_object(content, '$.hardwareInfo.supported64BitAbis') as supported64BitAbis
                  ,get_json_object(content, '$.hardwareInfo.supportedAbis') as supportedAbis
                  ,get_json_object(content, '$.hardwareInfo.tags') as tags
                  ,get_json_object(content, '$.hardwareInfo.time') as time
                  ,get_json_object(content, '$.hardwareInfo.type') as type
                  ,get_json_object(content, '$.hardwareInfo.user') as user
                  ,get_json_object(content, '$.hardwareInfo.cpuAbi') as cpuAbi
                  ,get_json_object(content, '$.hardwareInfo.cpuAbi2') as cpuAbi2
            from test_bigdata_data_hive_temp
        '''
        )
        
    col_toparse = ['bootloader', 'brand', 'display', 'externalStorageFree', 'externalStorageTotal', 'fingerprint', 'hardware', 'host', 'id', 'mainStorageFree'
                    , 'mainStorageTotal', 'manufacturer', 'model', 'product', 'ramTotalSize', 'release', 'serial', 'supported32BitAbis', 'supported64BitAbis'
                    , 'supportedAbis', 'tags', 'time', 'type', 'user', 'cpuAbi', 'cpuAbi2']
                    
    for col in col_toparse:
        hardwareInfo_df = hardwareInfo_df.withColumn(col, regexp_replace(col,'"',""))
    
    hardwareInfo_df = hardwareInfo_df.select(col_toparse + ['id_batch', 'dtime_datascore', 'part_date'])
    
    hardwareInfo_df.repartition(200).write.format('parquet').partitionBy('part_date').mode('append').saveAsTable('TEST_BIGDATA_OWNER.DATASCORE_HARDWARE_INFO')
    
    print('Execution time for hardwareInfo:', time.time() - start_time)
    print('')
    
    
def parse_callInfoList(whole_data_hive):
    
    start_time = time.time()
    print('Currently parsing callInfoList...')
    
    whole_data_hive.createOrReplaceTempView('test_bigdata_data_hive_temp')
    
    callInfoList_df = spark.sql(
        '''
            select
                   id_batch
                  ,dtime_datascore
                  ,part_date
                  ,get_json_object(content, '$.callInfoList[*].cachedPhotoID') as cachedPhotoID
                  ,get_json_object(content, '$.callInfoList[*].contactIdList') as contactIdList
                  ,get_json_object(content, '$.callInfoList[*].countryISO') as countryISO
                  ,get_json_object(content, '$.callInfoList[*].date') as date
                  ,get_json_object(content, '$.callInfoList[*].duration') as duration
                  ,get_json_object(content, '$.callInfoList[*].features') as features
                  ,get_json_object(content, '$.callInfoList[*].isRead') as isRead
                  ,get_json_object(content, '$.callInfoList[*].isVoiceEmailUri') as isVoiceEmailUri
                  ,get_json_object(content, '$.callInfoList[*].lastModified') as lastModified
                  ,get_json_object(content, '$.callInfoList[*].newCall') as newCall
                  ,get_json_object(content, '$.callInfoList[*].type') as type
                  ,get_json_object(content, '$.callInfoList[*].limitParamKey') as limitParamKey
            from test_bigdata_data_hive_temp
        '''
        )
        
    col_toparse = ['cachedPhotoID', 'contactIdList', 'countryISO', 'date', 'duration', 'features', 'isRead', 'isVoiceEmailUri', 'lastModified', 'newCall'
                    , 'type', 'limitParamKey']
    
    callInfoList_df = cleanse_and_parse_func(callInfoList_df, col_toparse)
    
    callInfoList_df.repartition(200).write.format('parquet').partitionBy('part_date').mode('append').saveAsTable('TEST_BIGDATA_OWNER.DATASCORE_CALL_INFO')
    
    print('Execution time for callInfoList:', time.time() - start_time)
    print('')
    
